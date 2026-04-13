import streamlit as st
import pandas as pd
from pathlib import Path
import io
from datetime import date
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
import numpy as np
import calendar
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

import requests

FILE_ID = "1FLoz9fyHlAke0MgrEgwSd8eTekH-zbCc"

@st.cache_data(ttl=3600, show_spinner="Загрузка с Google Drive...")
def load_from_gdrive(file_id: str) -> pd.DataFrame:
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return pd.read_excel(io.BytesIO(r.content), engine="openpyxl")

# ============ Page ============
st.set_page_config(page_title="Dashboard (Base)", layout="wide")

# ============ Excel export helper ============
HEADER_COLOR  = "1F4E79"
HEADER_FONT   = "FFFFFF"
ROW_ALT_COLOR = "DCE6F1"
FONT_NAME     = "Arial"

MONEY_COLS  = {"Выручка", "Сумма", "Средний чек", "Value", "PeakValue",
               "Итого за период", "Итого (сом)", "Остаток (сом)"}
INT_COLS    = {"Количество", "Чеков", "SKU_count", "Итого (кол)", "Остаток (шт)"}
PCT_COLS    = {"Share", "CumShare", "SKU_share", "Value_share", "Доля выручки"}
FLOAT2_COLS = {"Позиции/чек", "Товаров/чек", "Среднее/день (кол)", "Среднее/день (сом)"}

FMT_MONEY   = '#,##0'
FMT_INT     = '#,##0'
FMT_PCT     = '0.00%'
FMT_FLOAT2  = '0.00'
FMT_DATE    = 'DD.MM.YYYY'


def _col_fmt(col_name: str) -> str | None:
    if col_name in MONEY_COLS:  return FMT_MONEY
    if col_name in INT_COLS:    return FMT_INT
    if col_name in PCT_COLS:    return FMT_PCT
    if col_name in FLOAT2_COLS: return FMT_FLOAT2
    return None


def _auto_col_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 8), 55)


def df_to_sheet(ws, df: pd.DataFrame, sheet_title: str | None = None) -> None:
    start_row = 1
    if sheet_title:
        title_cell = ws.cell(row=1, column=1, value=sheet_title)
        title_cell.font = Font(name=FONT_NAME, bold=True, size=12, color=HEADER_COLOR)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
        start_row = 2

    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font = Font(name=FONT_NAME, bold=True, color=HEADER_FONT, size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=ci, value=str(col_name))
        cell.fill = header_fill; cell.font = header_font; cell.alignment = center

    ws.row_dimensions[start_row].height = 30
    alt_fill  = PatternFill("solid", fgColor=ROW_ALT_COLOR)
    data_font = Font(name=FONT_NAME, size=10)
    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(bottom=thin)

    for ri, row_data in enumerate(df.itertuples(index=False), start=1):
        excel_row = start_row + ri
        is_alt = (ri % 2 == 0)
        for ci, (col_name, val) in enumerate(zip(df.columns, row_data), start=1):
            cell = ws.cell(row=excel_row, column=ci)
            if isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime(); cell.number_format = FMT_DATE
            elif hasattr(val, "item"):
                cell.value = val.item()
            else:
                cell.value = val
            fmt = _col_fmt(col_name)
            if fmt: cell.number_format = fmt
            cell.font = data_font; cell.border = border
            if is_alt: cell.fill = alt_fill

    _auto_col_width(ws)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def build_excel_bytes(sheets: list[tuple[str, pd.DataFrame, str | None]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, df, sheet_title in sheets:
        safe_name = sheet_name[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=safe_name)
        df_to_sheet(ws, df.reset_index(drop=True), sheet_title)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def download_btn(label, sheets, filename="report.xlsx", key=None):
    if not sheets: return
    xlsx_bytes = build_excel_bytes(sheets)
    st.download_button(
        label=f"⬇️ {label}", data=xlsx_bytes, file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key,
    )


# ============ Data loading helpers ============
@st.cache_data(show_spinner=True)
def load_excel_from_bytes(xlsx_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    bio = io.BytesIO(xlsx_bytes)
    if sheet_name is None:
        xls = pd.ExcelFile(bio)
        preferred = ["база", "База", "Sheet1", "Лист1", "Лист 1"]
        sheet_name = next((s for s in preferred if s in xls.sheet_names), xls.sheet_names[0])
    df = pd.read_excel(bio, sheet_name=sheet_name, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df

@st.cache_data(show_spinner=True)
def load_excel_from_path(path: str, sheet_name: str | None = None) -> pd.DataFrame:
    if sheet_name is None:
        xls = pd.ExcelFile(path)
        preferred = ["база", "База", "Sheet1", "Лист1", "Лист 1"]
        sheet_name = next((s for s in preferred if s in xls.sheet_names), xls.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df

def basic_clean(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Дата" in df.columns:
        dt = pd.to_datetime(df["Дата"], errors="coerce", dayfirst=True)
        df["Дата"] = dt.dt.normalize()
        df = df[df["Дата"].notna()]
    for col in ["Количество", "Сумма", "Цена"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in ["Филиал", "Точки", "Номенклатура", "Категория", "Подкатегория", "Время"]:
        if col in df.columns:
            df[col] = df[col].astype("string").str.strip()
    return df

def validate_minimum(df: pd.DataFrame) -> None:
    required = ["Филиал", "Точки", "Номенклатура", "Количество", "Сумма", "Дата"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"В данных не хватает колонок: {missing}")
        st.stop()
    if df["Дата"].isna().all():
        st.error("Не удалось распознать ни одной даты в колонке 'Дата'. Проверь формат дат.")
        st.stop()

def safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0

def money(x: float) -> str:
    return f"{x:,.0f}".replace(",", " ")

# ============ Sidebar: source ============
st.sidebar.header("Источник данных")

import requests

source_mode = st.sidebar.radio(
    "Откуда брать данные?",
    ["Google Drive", "Локальный файл", "Загрузить вручную"],
    index=0,
)

GDRIVE_FILE_ID = "1FLoz9fyHlAke0MgrEgwSd8eTekH-zbCc"

@st.cache_data(ttl=3600, show_spinner="Загрузка с Google Sheets...")
def load_from_gdrive(file_id: str) -> pd.DataFrame:
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    bio = io.BytesIO(r.content)
    xls = pd.ExcelFile(bio)
    preferred = ["база", "База", "Sheet1", "Лист1", "Лист 1"]
    sheet = next((s for s in preferred if s in xls.sheet_names), xls.sheet_names[0])
    df = pd.read_excel(bio, sheet_name=sheet, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df

df = None

if source_mode == "Google Drive":
    try:
        df = load_from_gdrive(GDRIVE_FILE_ID)
        st.sidebar.caption(f"📄 Google Sheets (ID: ...{GDRIVE_FILE_ID[-6:]})")
    except Exception as e:
        st.error(f"Не удалось загрузить с Google Drive. Проверь доступ «по ссылке».\n\nОшибка: {e}")
        st.stop()

elif source_mode == "Локальный файл":
    p1 = Path("Итоговый_отчет1.xlsx")
    p0 = Path("Итоговый_отчет.xlsx")
    path = str(p1) if p1.exists() else str(p0)
    try:
        df = load_excel_from_path(path)
        st.sidebar.caption(f"Локальный файл: {path}")
    except Exception as e:
        st.error(f"Не удалось прочитать файл '{path}'. Ошибка: {e}")
        st.stop()

else:  # Загрузить вручную
    up = st.sidebar.file_uploader("Excel файл (.xlsx)", type=["xlsx"], key="uploader_xlsx")
    if up is not None:
        st.session_state["uploaded_xlsx_bytes"] = up.getvalue()
    if "uploaded_xlsx_bytes" in st.session_state:
        df = load_excel_from_bytes(st.session_state["uploaded_xlsx_bytes"])
df = basic_clean(df)
validate_minimum(df)

# ============ Sidebar: filters ============
min_ts = df["Дата"].min()
max_ts = df["Дата"].max()
min_d = min_ts.date()
max_d = max_ts.date()

branches_all = sorted(df["Филиал"].dropna().astype(str).unique().tolist())

if "applied_filters" not in st.session_state:
    st.session_state.applied_filters = {
        "date_range": (min_d, max_d),
        "branches": branches_all,
        "points": [],
        "categories": [],
        "subcategories": [],
        "abc_metric": "Сумма",
        "items": []
    }

if "filters_version" not in st.session_state:
    st.session_state.filters_version = 0

ap_from, ap_to = st.session_state.applied_filters["date_range"]
ap_branches = st.session_state.applied_filters.get("branches", branches_all)

need_reset = (ap_from < min_d) or (ap_to > max_d) or (not set(ap_branches).issubset(set(branches_all)))
if need_reset:
    st.session_state.applied_filters["date_range"] = (min_d, max_d)
    st.session_state.applied_filters["branches"] = branches_all
    st.session_state.applied_filters["points"] = []
    st.session_state.filters_version += 1

branches_selected_applied = st.session_state.applied_filters.get("branches", branches_all) or branches_all
df_for_points = df[df["Филиал"].isin(branches_selected_applied)]
points_all = sorted(df_for_points["Точки"].dropna().astype(str).unique().tolist())

ap_points = st.session_state.applied_filters.get("points", [])
if ap_points and not set(ap_points).issubset(set(points_all)):
    st.session_state.applied_filters["points"] = []
    st.session_state.filters_version += 1

st.sidebar.header("Фильтры")

with st.sidebar.form("filters_form", clear_on_submit=False):
    default_from, default_to = st.session_state.applied_filters["date_range"]
    default_branches = st.session_state.applied_filters.get("branches", branches_all)

    draft_date_range = st.date_input(
        "Период", value=(default_from, default_to),
        min_value=min_d, max_value=max_d, format="DD.MM.YYYY",
        key=f"date_range_input_{st.session_state.filters_version}",
    )

    draft_branches = st.multiselect(
        "Филиал", options=branches_all, default=default_branches,
        key=f"branches_input_{st.session_state.filters_version}",
    )

    branches_for_points = draft_branches or branches_all
    df_points_draft = df[df["Филиал"].isin(branches_for_points)]
    points_options = sorted(df_points_draft["Точки"].dropna().astype(str).unique().tolist())
    default_points = [p for p in st.session_state.applied_filters.get("points", []) if p in points_options]

    draft_points = st.multiselect(
        "Точки", options=points_options, default=default_points,
        key=f"points_input_{st.session_state.filters_version}",
    )

    df_cat_base = df[df["Филиал"].isin(branches_for_points)].copy()
    if draft_points:
        df_cat_base = df_cat_base[df_cat_base["Точки"].isin(draft_points)]

    categories_options = sorted(df_cat_base["Категория"].dropna().astype(str).unique().tolist())
    default_categories = [c for c in st.session_state.applied_filters.get("categories", []) if c in categories_options]

    draft_categories = st.multiselect(
        "Категория", options=categories_options, default=default_categories,
        key=f"categories_input_{st.session_state.filters_version}",
        help="Оставь пустым — будут показаны все категории",
    )

    df_sub_base = df_cat_base
    if draft_categories:
        df_sub_base = df_sub_base[df_sub_base["Категория"].isin(draft_categories)]

    subcategories_options = sorted(df_sub_base["Подкатегория"].dropna().astype(str).unique().tolist())
    default_subcategories = [sc for sc in st.session_state.applied_filters.get("subcategories", []) if sc in subcategories_options]

    draft_subcategories = st.multiselect(
        "Подкатегория", options=subcategories_options, default=default_subcategories,
        key=f"subcategories_input_{st.session_state.filters_version}",
    )

    df_item_base = df_cat_base
    if draft_categories:
        df_item_base = df_item_base[df_item_base["Категория"].isin(draft_categories)]
    if draft_subcategories:
        df_item_base = df_item_base[df_item_base["Подкатегория"].isin(draft_subcategories)]

    items_options = sorted(df_item_base["Номенклатура"].dropna().astype(str).unique().tolist())
    default_items = [x for x in st.session_state.applied_filters.get("items", []) if x in items_options]

    draft_items = st.multiselect(
        "Номенклатура", options=items_options, default=default_items,
        key=f"items_input_{st.session_state.filters_version}",
        help="Оставь пустым — будут показаны все номенклатуры.",
    )

    draft_metric = st.radio(
        "ABC метрика", options=["Сумма", "Количество"], index=0,
        horizontal=True, key=f"abc_metric_input_{st.session_state.filters_version}",
    )

    apply_btn = st.form_submit_button("Применить")

if apply_btn:
    if isinstance(draft_date_range, tuple) and len(draft_date_range) == 2:
        d_from, d_to = draft_date_range
    else:
        d_from = d_to = draft_date_range
    if d_from > d_to:
        d_from, d_to = d_to, d_from
    if not draft_branches:
        draft_branches = branches_all

    st.session_state.applied_filters.update({
        "date_range": (d_from, d_to),
        "branches": draft_branches,
        "points": draft_points,
        "categories": draft_categories,
        "subcategories": draft_subcategories,
        "abc_metric": draft_metric,
        "items": draft_items,
    })

# ---- применяем фильтры ----
d_from, d_to = st.session_state.applied_filters["date_range"]
branches_selected = st.session_state.applied_filters.get("branches", branches_all) or branches_all
points_selected   = st.session_state.applied_filters.get("points", [])

from_ts = pd.Timestamp(d_from)
to_ts   = pd.Timestamp(d_to)

df_filtered = df[(df["Дата"] >= from_ts) & (df["Дата"] <= to_ts)].copy()
df_filtered = df_filtered[df_filtered["Филиал"].isin(branches_selected)].copy()
if points_selected:
    df_filtered = df_filtered[df_filtered["Точки"].isin(points_selected)].copy()

categories_selected    = st.session_state.applied_filters.get("categories", [])
subcategories_selected = st.session_state.applied_filters.get("subcategories", [])
if categories_selected:
    df_filtered = df_filtered[df_filtered["Категория"].isin(categories_selected)].copy()
if subcategories_selected:
    df_filtered = df_filtered[df_filtered["Подкатегория"].isin(subcategories_selected)].copy()

items_selected = st.session_state.applied_filters.get("items", [])
if items_selected:
    df_filtered = df_filtered[df_filtered["Номенклатура"].isin(items_selected)].copy()

# ============ KPI helpers ============
checks_col = "Склад/Товар"

def count_checks(frame: pd.DataFrame) -> int:
    if checks_col not in frame.columns or frame.empty: return 0
    s = (frame[checks_col].astype(str).str.strip()
         .str.replace(r"\s+", " ", regex=True)
         .replace({"": pd.NA, "nan": pd.NA}).dropna())
    return int(s.nunique())

def kpi_table(frame: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=group_cols + ["Выручка","Количество","Чеков","Средний чек","Позиции/чек","Товаров/чек"])
    g = frame.groupby(group_cols, dropna=False).agg(
        Выручка=("Сумма", "sum"), Количество=("Количество", "sum"), Строк=("Сумма", "size"),
    ).reset_index()
    checks = frame[group_cols + [checks_col]].copy()
    checks[checks_col] = (checks[checks_col].astype(str).str.strip()
                          .str.replace(r"\s+", " ", regex=True)
                          .replace({"": pd.NA, "nan": pd.NA}))
    checks = checks.dropna(subset=[checks_col])
    if not checks.empty:
        checks_cnt = checks.groupby(group_cols)[checks_col].nunique().reset_index().rename(columns={checks_col: "Чеков"})
        g = g.merge(checks_cnt, on=group_cols, how="left")
    else:
        g["Чеков"] = 0
    g["Чеков"] = g["Чеков"].fillna(0).astype(int)
    g["Средний чек"]  = g.apply(lambda r: r["Выручка"]/r["Чеков"] if r["Чеков"] else 0.0, axis=1)
    g["Позиции/чек"]  = g.apply(lambda r: r["Строк"]/r["Чеков"]   if r["Чеков"] else 0.0, axis=1)
    g["Товаров/чек"]  = g.apply(lambda r: r["Количество"]/r["Чеков"] if r["Чеков"] else 0.0, axis=1)
    total_sales = float(g["Выручка"].sum()) or 0.0
    g["Доля выручки"] = g["Выручка"] / total_sales if total_sales else 0.0
    return g.sort_values("Выручка", ascending=False).drop(columns=["Строк"])

kpi_branch       = kpi_table(df_filtered, ["Филиал"])
kpi_branch_point = kpi_table(df_filtered, ["Филиал", "Точки"])

# ============ ABC ============
metric     = st.session_state.applied_filters.get("abc_metric", "Сумма")
metric_col = "Сумма" if metric == "Сумма" else "Количество"
A_thr, B_thr = 0.80, 0.95

def build_abc(df_in, group_cols):
    g = (df_in.groupby(group_cols, dropna=False)[metric_col]
         .sum().reset_index().rename(columns={metric_col: "Value"})
         .sort_values("Value", ascending=False).reset_index(drop=True))
    total = float(g["Value"].sum()) if not g.empty else 0.0
    if total <= 0 or pd.isna(total):
        g["Share"] = g["CumShare"] = 0.0
    else:
        g["Share"] = g["Value"] / total
        g["CumShare"] = g["Share"].cumsum()
    g["ABC"] = g["CumShare"].apply(lambda x: "A" if x <= A_thr else ("B" if x <= B_thr else "C"))
    return g

def abc_summary(abc_df):
    if abc_df.empty:
        return pd.DataFrame(columns=["ABC","SKU_count","SKU_share","Value","Value_share"])
    total_sku = len(abc_df)
    total_val = float(abc_df["Value"].sum()) or 0.0
    s = abc_df.groupby("ABC")["Value"].agg(SKU_count="count", Value="sum").reset_index()
    s["SKU_share"]   = s["SKU_count"] / total_sku if total_sku else 0.0
    s["Value_share"] = s["Value"] / total_val if total_val else 0.0
    s["ABC"] = pd.Categorical(s["ABC"], categories=["A","B","C"], ordered=True)
    return s.sort_values("ABC").reset_index(drop=True)

def pareto_chart(abc_df, label_col, top_n=30):
    d = abc_df.head(top_n).copy()
    fig, ax1 = plt.subplots(figsize=(12, 5))
    ax1.bar(range(len(d)), d["Value"])
    ax1.set_ylabel(metric_col); ax1.set_xlabel(label_col)
    ax1.set_xticks(range(len(d)))
    ax1.set_xticklabels(d[label_col].astype(str).tolist(), rotation=75, ha="right", fontsize=8)
    ax2 = ax1.twinx()
    ax2.plot(range(len(d)), d["CumShare"].values, marker="o")
    ax2.set_ylabel("Кумулятивная доля"); ax2.set_ylim(0, 1.05)
    ax2.axhline(A_thr, linestyle="--"); ax2.axhline(B_thr, linestyle="--")
    ax1.set_title(f"Pareto (Top {top_n})")
    fig.tight_layout()
    return fig

abc_overall   = build_abc(df_filtered, ["Номенклатура"])
abc_by_branch = build_abc(df_filtered, ["Филиал", "Номенклатура"])
abc_stats     = abc_summary(abc_overall)

# ============ Time analysis ============
df_time = df_filtered.copy()
if "Время" in df_time.columns:
    t = pd.to_datetime(df_time["Время"].astype(str).str.strip(), errors="coerce")
    df_time["Hour"] = t.dt.hour
    df_time = df_time[df_time["Hour"].notna()].copy()
    df_time["Hour"] = df_time["Hour"].astype(int)
else:
    df_time = df_time.iloc[0:0].copy()

if not df_time.empty:
    by_branch_hour = (df_time.groupby(["Филиал","Hour"], dropna=False)[metric_col]
                      .sum().reset_index().rename(columns={metric_col: "Value"}))
    pivot_branch_hour = (by_branch_hour.pivot_table(index="Филиал", columns="Hour", values="Value", fill_value=0)
                         .sort_index(axis=1))
    peak_by_branch = (by_branch_hour.sort_values(["Филиал","Value"], ascending=[True,False])
                      .groupby("Филиал", as_index=False).head(1)
                      .rename(columns={"Hour":"PeakHour","Value":"PeakValue"}))
else:
    by_branch_hour    = pd.DataFrame(columns=["Филиал","Hour","Value"])
    pivot_branch_hour = pd.DataFrame()
    peak_by_branch    = pd.DataFrame(columns=["Филиал","PeakHour","PeakValue"])

# ============ Trend ============
def aggregate_for_chart(df_in, metric_col, freq):
    if df_in.empty: return pd.DataFrame(columns=["Period","Value"])
    return (df_in.set_index("Дата")[metric_col].resample(freq).sum().reset_index()
            .rename(columns={"Дата":"Period", metric_col:"Value"}))

def pick_freq(d_from, d_to):
    days = (d_to - d_from).days + 1
    if days <= 62:  return "D"
    if days <= 370: return "W-MON"
    return "MS"

freq       = pick_freq(d_from, d_to)
cur_series = aggregate_for_chart(df_filtered, metric_col, freq)
cur_total  = float(df_filtered[metric_col].sum()) if not df_filtered.empty else 0.0

df_daily_cur = (df_filtered.set_index("Дата")[metric_col].resample("D").sum().reset_index()
                .rename(columns={"Дата":"Day", metric_col:"Value"}))

best_day = best_val = None
if not df_daily_cur.empty:
    idx = df_daily_cur["Value"].idxmax()
    best_day = df_daily_cur.loc[idx, "Day"].date()
    best_val = float(df_daily_cur.loc[idx, "Value"])


# ============================================================
# TAB 6 helper — calendar heatmap
# ============================================================
def build_calendar_heatmap(daily_series: pd.Series, year: int, month: int,
                            metric_label: str, vmax: float, cmap) -> plt.Figure:
    cal = calendar.monthcalendar(year, month)
    n_weeks = len(cal)
    day_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    month_ru  = ["", "Январь","Февраль","Март","Апрель","Май","Июнь",
                 "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

    cell_size = 1.0
    pad       = 0.06
    fig_w     = 7 * cell_size + 1.2
    fig_h     = (n_weeks + 1.2) * cell_size

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, 7)
    ax.set_ylim(0, n_weeks + 1)
    ax.axis("off")
    ax.set_aspect("equal")
    fig.patch.set_facecolor("#F7F9FC")

    ax.text(3.5, n_weeks + 0.65, f"{month_ru[month]} {year}",
            ha="center", va="center", fontsize=13, fontweight="bold", color="#1F3864")

    for col, dname in enumerate(day_names):
        color = "#C0392B" if col >= 5 else "#1F3864"
        ax.text(col + 0.5, n_weeks + 0.15, dname,
                ha="center", va="center", fontsize=9, fontweight="bold", color=color)

    norm = mcolors.Normalize(vmin=0, vmax=max(vmax, 1))

    for week_idx, week in enumerate(cal):
        row = n_weeks - 1 - week_idx
        for dow, day_num in enumerate(week):
            x = dow; y = row
            if day_num == 0:
                rect = mpatches.FancyBboxPatch(
                    (x + pad, y + pad), cell_size - 2*pad, cell_size - 2*pad,
                    boxstyle="round,pad=0.04", linewidth=0,
                    facecolor="#ECEFF4", zorder=1)
                ax.add_patch(rect)
                continue

            d = date(year, month, day_num)
            val = float(daily_series.get(d, 0.0))
            is_weekend = (dow >= 5)

            if val > 0:
                rgba = cmap(norm(val))
                face  = rgba
                tcolor = "white" if norm(val) > 0.55 else "#1F3864"
            else:
                face   = "#ECEFF4" if not is_weekend else "#FAE5E5"
                tcolor = "#AABBCC"

            rect = mpatches.FancyBboxPatch(
                (x + pad, y + pad), cell_size - 2*pad, cell_size - 2*pad,
                boxstyle="round,pad=0.05", linewidth=0.5,
                edgecolor="#CFD8E3", facecolor=face, zorder=1)
            ax.add_patch(rect)

            ax.text(x + 0.12, y + cell_size - 0.17, str(day_num),
                    ha="left", va="top", fontsize=7, color=tcolor,
                    alpha=0.7, zorder=2)

            if val > 0:
                val_str = f"{val:,.1f}".rstrip("0").rstrip(".") if val != int(val) else str(int(val))
                ax.text(x + 0.5, y + 0.42, val_str,
                        ha="center", va="center", fontsize=10.5,
                        fontweight="bold", color=tcolor, zorder=2)

    fig.tight_layout(pad=0.3)
    return fig


def _get_months_for_period(d_from, d_to):
    months = []
    cur_d = date(d_from.year, d_from.month, 1)
    end_d = date(d_to.year, d_to.month, 1)
    while cur_d <= end_d:
        months.append((cur_d.year, cur_d.month))
        if cur_d.month == 12:
            cur_d = date(cur_d.year + 1, 1, 1)
        else:
            cur_d = date(cur_d.year, cur_d.month + 1, 1)
    return months


def _render_one_panel(df_source, all_items, cal_metric, show_all_months,
                      panel_idx, panel_colors):
    color_accent = panel_colors[panel_idx % 3]

    st.markdown(
        f"<div style='background:{color_accent};border-radius:6px;"
        f"padding:4px 12px;margin-bottom:6px;'>"
        f"<span style='color:white;font-weight:700;font-size:13px;'>"
        f"Позиция {panel_idx + 1}</span></div>",
        unsafe_allow_html=True,
    )

    chosen_item = st.selectbox(
        "Номенклатура",
        all_items,
        key=f"cal_item_{panel_idx}",
        label_visibility="collapsed",
    )

    df_item = df_source[df_source["Номенклатура"].astype(str) == str(chosen_item)].copy()

    if df_item.empty:
        st.warning(f"Нет продаж")
        return

    daily = df_item.groupby(df_item["Дата"].dt.date)[cal_metric].sum()
    daily.index = pd.to_datetime(daily.index).date

    total_val = float(daily.sum())
    days_sold = int((daily > 0).sum())
    avg_day   = total_val / days_sold if days_sold else 0.0
    label_val = "Продано" if cal_metric == "Количество" else "Выручка"

    st.markdown(
        f"<div style='background:#F0F4FA;border-left:4px solid {color_accent};"
        f"border-radius:4px;padding:8px 10px;margin:4px 0 8px 0;font-size:12px;color:#555;'>"
        f"<b style='color:{color_accent};'>{chosen_item}</b><br>"
        f"<span>📦 Всего ({label_val}): <b>"
        + (f"{total_val:,.2f}".rstrip("0").rstrip(".") if cal_metric == "Количество" else money(total_val))
        + f"</b></span>&nbsp;&nbsp;"
        f"<span>📅 Дней с продажами: <b>{days_sold}</b></span>&nbsp;&nbsp;"
        f"<span>⌀ В среднем/день: <b>"
        + (f"{avg_day:,.2f}".rstrip("0").rstrip(".") if cal_metric == "Количество" else money(avg_day))
        + "</b></span>"
        + (
            f"&nbsp;&nbsp;<span>🏆 Пик: <b>{daily.idxmax():%d.%m.%Y} ({daily[daily.idxmax()]:g})</b></span>"
            if not daily.empty else ""
        )
        + "</div>",
        unsafe_allow_html=True,
    )

    cmaps_qty = ["YlOrRd", "Blues", "Greens", "Purples", "Oranges", "RdPu"]
    cmaps_sum = ["Blues",  "Purples", "YlOrBr", "BuGn", "PuRd", "GnBu"]
    ci = panel_idx % 6
    cmap = plt.get_cmap(cmaps_qty[ci] if cal_metric == "Количество" else cmaps_sum[ci])

    vmax = float(daily.max()) if not daily.empty else 1.0

    all_months = _get_months_for_period(d_from, d_to)

    if not show_all_months:
        month_names = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
                       "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
        sel_ym = st.selectbox(
            "Месяц",
            options=all_months,
            format_func=lambda ym: f"{month_names[ym[1]]} {ym[0]}",
            key=f"cal_month_{panel_idx}",
        )
        draw_months = [sel_ym]
    else:
        draw_months = all_months

    for year, month in draw_months:
        fig = build_calendar_heatmap(daily, year, month, cal_metric, vmax, cmap)
        st.pyplot(fig, clear_figure=True)

    st.caption(f"Цвет: 0 → {vmax:g} | {cal_metric}")


def calendar_heatmap_section(df_source: pd.DataFrame, metric_col: str):
    if df_source.empty:
        st.info("Нет данных по выбранным фильтрам.")
        return

    all_items = sorted(df_source["Номенклатура"].dropna().astype(str).unique().tolist())
    if not all_items:
        st.info("Нет доступных позиций номенклатуры.")
        return

    if "cal_group_count" not in st.session_state:
        st.session_state["cal_group_count"] = 1

    ctrl1, ctrl2 = st.columns([2, 2])
    with ctrl1:
        cal_metric = st.radio(
            "Метрика для всех панелей", ["Количество", "Сумма"],
            key="cal_metric_radio", horizontal=True
        )
    with ctrl2:
        show_all_months = st.checkbox(
            "Показывать все месяцы", value=True, key="cal_all_months"
        )

    BASE_COLORS = [
        ["#C0392B", "#2471A3", "#1E8449"],
        ["#7D3C98", "#D35400", "#117A65"],
        ["#1A5276", "#B7950B", "#922B21"],
        ["#0E6655", "#6E2F9A", "#C0392B"],
    ]

    n_groups = st.session_state["cal_group_count"]

    for g in range(n_groups):
        colors = BASE_COLORS[g % len(BASE_COLORS)]

        if g == 0:
            st.divider()
        else:
            st.markdown(
                f"<div style='display:flex;align-items:center;margin:18px 0 6px 0;'>"
                f"<div style='flex:1;height:2px;background:linear-gradient(90deg,"
                f"{colors[0]}44,{colors[1]}44,{colors[2]}44);border-radius:2px;'></div>"
                f"<span style='margin:0 12px;font-size:12px;font-weight:600;"
                f"color:#888;white-space:nowrap;'>Блок сравнения {g + 1}</span>"
                f"<div style='flex:1;height:2px;background:linear-gradient(90deg,"
                f"{colors[2]}44,{colors[1]}44,{colors[0]}44);border-radius:2px;'></div>"
                f"</div>",
                unsafe_allow_html=True,
            )

        panel_cols = st.columns(3, gap="medium")
        for i, col in enumerate(panel_cols):
            with col:
                panel_idx = g * 3 + i
                _render_one_panel(
                    df_source, all_items, cal_metric,
                    show_all_months, panel_idx, colors
                )

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 6])

    with btn_col1:
        if st.button(
            "＋  Добавить блок",
            key="cal_add_group",
            help="Добавить ещё один ряд из 3 панелей для сравнения",
            use_container_width=True,
        ):
            st.session_state["cal_group_count"] += 1
            st.rerun()

    with btn_col2:
        if n_groups > 1:
            if st.button(
                "－  Убрать блок",
                key="cal_remove_group",
                help="Удалить последний блок",
                use_container_width=True,
            ):
                st.session_state["cal_group_count"] -= 1
                removed_base = (n_groups - 1) * 3
                for i in range(3):
                    for k in [f"cal_item_{removed_base+i}", f"cal_month_{removed_base+i}"]:
                        st.session_state.pop(k, None)
                st.rerun()


# ============================================================
# BOM (разбивка наборов) loader
# ============================================================
SEASONAL_KEYWORDS = [
    '23 февраля', '14 февраля', '8 марта', '1 сентября',
    'рамадан', 'новый год', 'пасх', 'наурыз', 'весна',
]

@st.cache_data(show_spinner=False)
def load_bom() -> pd.DataFrame:
    """Загружает разбивку наборов. Возвращает пустой DF если файл не найден."""
    bom_path = Path("разбивка_наборов.xlsx")
    if not bom_path.exists():
        return pd.DataFrame(columns=["Набор", "Компонент", "Кол"])
    df = pd.read_excel(str(bom_path))
    df.columns = ["Набор", "Компонент", "Кол"]
    df["Набор"]     = df["Набор"].astype(str).str.strip()
    df["Компонент"] = df["Компонент"].astype(str).str.strip()
    df["Кол"]       = pd.to_numeric(df["Кол"], errors="coerce").fillna(0)
    # исключаем сезонные
    def _seasonal(name):
        n = name.lower()
        return any(k in n for k in SEASONAL_KEYWORDS)
    df = df[~df["Набор"].apply(_seasonal)].reset_index(drop=True)
    return df


def build_components_breakdown(
    df_sales: pd.DataFrame,
    bom: pd.DataFrame,
    days_in_period: int,
    coverage_days: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Возвращает два DataFrame:
      1. Компоненты — суммарная потребность (прямые продажи + из наборов)
      2. Наборы     — сток укомплектованных наборов

    Единицы:
      - Для наборов с коэффициентом < 1 (драже) — компонент в кг
      - Для остальных — в штуках
    """
    if df_sales.empty or bom.empty:
        return pd.DataFrame(), pd.DataFrame()

    # ── агрегат продаж по номенклатуре ─────────────────────
    sales_agg = (df_sales.groupby("Номенклатура", dropna=False)["Количество"]
                 .sum().reset_index()
                 .rename(columns={"Номенклатура": "Позиция", "Количество": "Прямые продажи"}))
    sales_map = sales_agg.set_index("Позиция")["Прямые продажи"].to_dict()

    # ── набор-позиции которые продавались ──────────────────
    sets_in_bom = set(bom["Набор"].unique())
    sets_sold   = {k: v for k, v in sales_map.items() if k in sets_in_bom}

    # ─── ТАБЛИЦА 1: КОМПОНЕНТЫ ─────────────────────────────
    # Прямые продажи компонентов (которые НЕ являются наборами)
    direct = {k: v for k, v in sales_map.items() if k not in sets_in_bom}

    # Из наборов
    from_sets: dict[str, float] = {}
    for _, row in bom.iterrows():
        набор     = row["Набор"]
        компонент = row["Компонент"]
        кол       = float(row["Кол"])
        продано   = sets_sold.get(набор, 0)
        from_sets[компонент] = from_sets.get(компонент, 0) + продано * кол

    # Все компоненты
    all_components = set(direct) | set(from_sets)

    # Определяем единицу: если в bom коэффициент < 1 → кг, иначе → шт
    comp_unit: dict[str, str] = {}
    for _, row in bom.iterrows():
        c = row["Компонент"]
        if c not in comp_unit:
            comp_unit[c] = "кг" if float(row["Кол"]) < 1 else "шт"

    rows_comp = []
    for comp in sorted(all_components):
        direct_qty   = direct.get(comp, 0)
        from_sets_qty = from_sets.get(comp, 0)
        total        = direct_qty + from_sets_qty
        avg_day      = total / days_in_period if days_in_period else 0
        unit         = comp_unit.get(comp, "шт")
        if unit == "кг":
            stock = round(avg_day * coverage_days, 2)
        else:
            stock = max(int(np.ceil(avg_day * coverage_days)), 1) if total > 0 else 0
        rows_comp.append({
            "Компонент":        comp,
            "Ед":               unit,
            "Прямые продажи":   round(direct_qty, 3) if unit == "кг" else int(direct_qty),
            "Из наборов":       round(from_sets_qty, 3) if unit == "кг" else int(from_sets_qty),
            "Итого":            round(total, 3) if unit == "кг" else int(total),
            "Среднее/день":     round(avg_day, 3),
            f"Остаток ({unit})":   stock,
        })

    df_comp = pd.DataFrame(rows_comp)
    if not df_comp.empty:
        df_comp = df_comp.sort_values(f"Сток (шт)" if "Сток (шт)" in df_comp.columns else df_comp.columns[-1],
                                      ascending=False).reset_index(drop=True)

    # ─── ТАБЛИЦА 2: НАБОРЫ ─────────────────────────────────
    rows_sets = []
    for набор in sorted(sets_in_bom):
        продано  = sets_sold.get(набор, 0)
        avg_day  = продано / days_in_period if days_in_period else 0
        stock    = max(int(np.ceil(avg_day * coverage_days)), 1) if продано > 0 else 0
        rows_sets.append({
            "Набор":                набор,
            "Продано (шт)":         int(продано),
            "Среднее/день":         round(avg_day, 2),
            "Остаток (коробок)":       stock,
        })

    df_sets = pd.DataFrame(rows_sets).sort_values("Продано (шт)", ascending=False).reset_index(drop=True)
    return df_comp, df_sets


# ============================================================
# TAB 7 helper — Safety Stock
# ============================================================
def build_safety_stock(df_in: pd.DataFrame, days_in_period: int, coverage_days: int) -> pd.DataFrame:
    """
    Считает безопасный сток по номенклатуре суммарно по всем филиалам.

    Формула:
        Среднее/день = Итого за период / days_in_period
        Безопасный сток = ceil(Среднее/день * coverage_days), минимум 1

    Возвращает колонки:
        Номенклатура, [Категория], [Подкатегория],
        Итого (кол), Итого (сом),
        Среднее/день (кол), Среднее/день (сом),
        Остаток (шт), Остаток (сом),
        ABC
    """
    if df_in.empty:
        return pd.DataFrame()

    extra_cols = [c for c in ["Категория", "Подкатегория"] if c in df_in.columns]

    # агрегат по номенклатуре
    g = (df_in.groupby("Номенклатура", dropna=False)
         .agg(Итого_кол=("Количество", "sum"), Итого_сом=("Сумма", "sum"))
         .reset_index())

    g["Среднее/день (кол)"] = g["Итого_кол"] / days_in_period
    g["Среднее/день (сом)"] = g["Итого_сом"] / days_in_period
    g["Остаток (шт)"] = (
        (g["Среднее/день (кол)"] * coverage_days)
        .apply(lambda x: max(int(np.ceil(x)), 1))
    )
    g["Остаток (сом)"] = g["Среднее/день (сом)"] * coverage_days

    # ABC из текущего расчёта
    if not abc_overall.empty:
        abc_map = abc_overall.set_index("Номенклатура")["ABC"].to_dict()
        g["ABC"] = g["Номенклатура"].map(abc_map).fillna("—")
    else:
        g["ABC"] = "—"

    # мета (категория / подкатегория)
    if extra_cols:
        meta = (df_in[["Номенклатура"] + extra_cols]
                .drop_duplicates(subset=["Номенклатура"])
                .reset_index(drop=True))
        g = g.merge(meta, on="Номенклатура", how="left")

    g = g.rename(columns={"Итого_кол": "Итого (кол)", "Итого_сом": "Итого (сом)"})

    col_order = (["Номенклатура"] + extra_cols +
                 ["Итого (кол)", "Итого (сом)",
                  "Среднее/день (кол)", "Среднее/день (сом)",
                  "Остаток (шт)", "Остаток (сом)",
                  "ABC"])
    g = g[[c for c in col_order if c in g.columns]]
    return g.sort_values("Остаток (шт)", ascending=False).reset_index(drop=True)


# ============ UI ============
min_year = int(df_filtered["Дата"].dt.year.min()) if not df_filtered.empty else int(pd.Timestamp.today().year)
max_year = int(df_filtered["Дата"].dt.year.max()) if not df_filtered.empty else int(pd.Timestamp.today().year)

sel_branches = st.session_state.applied_filters.get("branches", [])
sel_points   = st.session_state.applied_filters.get("points", [])

st.title(f"Sales Dashboard — {min_year}–{max_year}")
st.caption(
    f"Текущий срез: {d_from:%d.%m.%Y} — {d_to:%d.%m.%Y} | "
    f"Филиалы: {len(sel_branches)} | Точки: {len(sel_points)} | "
    f"Метрика: {metric_col}"
)

sales  = float(df_filtered["Сумма"].sum())      if "Сумма"      in df_filtered.columns else 0.0
qty    = float(df_filtered["Количество"].sum()) if "Количество" in df_filtered.columns else 0.0

if checks_col in df_filtered.columns:
    checks_series = (df_filtered[checks_col].astype(str).str.strip()
                     .str.replace(r"\s+", " ", regex=True))
    checks_cnt = int(checks_series.replace({"": pd.NA, "nan": pd.NA}).dropna().nunique())
else:
    checks_cnt = 0

avg_check = safe_div(sales, checks_cnt)

st.markdown("### KPI")
c1, c2, c3, c4 = st.columns(4)
c1.metric("Выручка",     money(sales))
c2.metric("Количество",  f"{qty:,.0f}".replace(",", " "))
c3.metric("Чеков",       f"{checks_cnt:,}".replace(",", " "))
c4.metric("Средний чек", money(avg_check))

days_cnt     = (d_to - d_from).days + 1
main_total   = float(df_filtered[metric_col].sum()) if metric_col in df_filtered.columns else 0.0
main_per_day = safe_div(main_total, days_cnt)
c4.metric(f"{metric_col} / день", money(main_per_day))

st.divider()
with st.expander("⬇️ Скачать все таблицы одним файлом"):
    if not df_filtered.empty:
        all_sheets: list[tuple[str, pd.DataFrame, str | None]] = [
            ("KPI Филиалы",     kpi_branch,       "KPI по филиалам"),
            ("KPI Точки",       kpi_branch_point, "KPI по точкам"),
            ("ABC Сводка",      abc_stats,        "ABC сводка (A/B/C)"),
            ("ABC Общий",       abc_overall,      f"ABC по номенклатуре — {metric_col}"),
            ("ABC по филиалам", abc_by_branch,    "ABC по номенклатуре и филиалам"),
        ]
        if not peak_by_branch.empty:
            all_sheets.append(("Пиковые часы", peak_by_branch, "Пиковый час по каждому филиалу"))
        if not df_daily_cur.empty:
            top10_export = df_daily_cur.sort_values("Value", ascending=False).head(10).copy()
            top10_export["Day"] = pd.to_datetime(top10_export["Day"])
            all_sheets.append(("Топ-10 дней", top10_export, "Топ-10 дней (пики выручки)"))
        fname = f"dashboard_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx"
        download_btn("Скачать все таблицы", all_sheets, filename=fname, key="dl_all")
    else:
        st.info("Нет данных для экспорта.")

st.divider()

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "ABC (общий)", "ABC по филиалам",
    "Время (пики кассы)", "Тренд", "KPI по филиалам",
    "📅 Продажи по дням",
    "🏭 Остатки на складе",
    "📋 План / Факт по филиалам",
])

# ============================================================
# TAB 1 — ABC общий
# ============================================================
with tab1:
    st.subheader(f"ABC по номенклатуре (общий) — метрика: {metric_col}")
    c1, c2 = st.columns([1, 2])
    with c1:
        st.markdown("**Сводка A/B/C**")
        st.dataframe(abc_stats, use_container_width=True, hide_index=True)
        download_btn("Скачать сводку A/B/C",
                     [("ABC Сводка", abc_stats, "ABC сводка (A/B/C)")],
                     filename="abc_summary.xlsx", key="dl_abc_stats")
        top_n = st.slider("Top-N для Pareto", 10, 200, 30, 10)
        st.caption("Линии: 80% (A) и 95% (B).")
    with c2:
        if abc_overall.empty:
            st.info("Нет данных для Pareto по выбранным фильтрам.")
        else:
            st.pyplot(pareto_chart(abc_overall, label_col="Номенклатура", top_n=top_n), clear_figure=True)
    st.markdown("**Таблица ABC**")
    st.dataframe(abc_overall.head(500), use_container_width=True)
    download_btn("Скачать таблицу ABC (общий)",
                 [("ABC Общий", abc_overall, f"ABC по номенклатуре — {metric_col}"),
                  ("ABC Сводка", abc_stats,  "ABC сводка (A/B/C)")],
                 filename=f"abc_overall_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_abc_overall")

# ============================================================
# TAB 2 — ABC по филиалам
# ============================================================
with tab2:
    st.subheader("ABC по номенклатуре в разрезе филиалов")
    branches_in_data = sorted(abc_by_branch["Филиал"].dropna().astype(str).unique().tolist())
    if branches_in_data:
        selected_branch = st.selectbox("Филиал для Pareto", branches_in_data)
        abc_one_branch  = abc_by_branch[abc_by_branch["Филиал"].astype(str) == str(selected_branch)].copy()
        abc_one_branch  = abc_one_branch.sort_values("Value", ascending=False).reset_index(drop=True)
        total_b = float(abc_one_branch["Value"].sum()) if not abc_one_branch.empty else 0.0
        if total_b > 0:
            abc_one_branch["Share"]    = abc_one_branch["Value"] / total_b
            abc_one_branch["CumShare"] = abc_one_branch["Share"].cumsum()
            abc_one_branch["ABC"]      = abc_one_branch["CumShare"].apply(
                lambda x: "A" if x <= A_thr else ("B" if x <= B_thr else "C"))
        else:
            abc_one_branch[["Share","CumShare"]] = 0.0
            abc_one_branch["ABC"] = "C"
        top_n_b = st.slider("Top-N для Pareto (филиал)", 10, 200, 30, 10, key="topn_branch")
        if not abc_one_branch.empty:
            st.pyplot(pareto_chart(abc_one_branch, label_col="Номенклатура", top_n=top_n_b), clear_figure=True)
        st.markdown("**Таблица ABC (выбранный филиал)**")
        st.dataframe(abc_one_branch.head(500), use_container_width=True)
        safe_br = selected_branch.replace(" ", "_")[:20]
        download_btn(f"Скачать ABC — {selected_branch}",
                     [("ABC "+selected_branch[:25], abc_one_branch, f"ABC: {selected_branch}")],
                     filename=f"abc_{safe_br}_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_abc_one_branch")
    st.markdown("**Полная таблица ABC по филиалам**")
    st.dataframe(abc_by_branch.head(500), use_container_width=True)
    download_btn("Скачать полную таблицу ABC по филиалам",
                 [("ABC по филиалам", abc_by_branch, "ABC по номенклатуре и филиалам")],
                 filename=f"abc_by_branch_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_abc_by_branch")

# ============================================================
# TAB 3 — Время (пики кассы)
# ============================================================
with tab3:
    st.subheader(f"Касса по времени (метрика: {metric_col})")
    if df_time.empty:
        st.info("Нет корректных данных в колонке 'Время' для выбранных фильтров.")
    else:
        c1, c2 = st.columns([1, 2])
        with c1:
            st.markdown("**Пиковый час по каждому филиалу**")
            view = peak_by_branch.copy()
            if not view.empty:
                view["PeakHour"] = view["PeakHour"].apply(lambda h: f"{int(h):02d}:00")
            st.dataframe(view, use_container_width=True, hide_index=True)
            download_btn("Скачать пиковые часы",
                         [("Пиковые часы", peak_by_branch, "Пиковый час по каждому филиалу")],
                         filename=f"peak_hours_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_peak_hours")
        with c2:
            st.markdown("**Матрица: филиалы × часы**")
            st.dataframe(pivot_branch_hour, use_container_width=True)
            pivot_export = pivot_branch_hour.reset_index()
            pivot_export.columns = [str(c) for c in pivot_export.columns]
            download_btn("Скачать матрицу часов",
                         [("Матрица часов", pivot_export, f"Филиалы × Часы — {metric_col}")],
                         filename=f"hour_matrix_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_hour_matrix")
        branches_h = sorted(by_branch_hour["Филиал"].dropna().astype(str).unique().tolist())
        sel = st.selectbox("Показать график по филиалу", branches_h)
        d_bh = by_branch_hour[by_branch_hour["Филиал"].astype(str) == str(sel)].sort_values("Hour")
        st.line_chart(d_bh.set_index("Hour")["Value"])
        download_btn("Скачать детальную таблицу (все часы × все филиалы)",
                     [("Часы по филиалам", by_branch_hour, f"Выручка/Количество по часам — {metric_col}")],
                     filename=f"by_branch_hour_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_by_branch_hour")

# ============================================================
# TAB 4 — Тренд
# ============================================================
with tab4:
    st.subheader(f"Тренд за период — метрика: {metric_col}")
    compare_on = st.checkbox("Сравнить с другим периодом", value=False)
    comp_from = comp_to = None
    comp_series = pd.DataFrame(columns=["Period","Value"])
    comp_total = 0.0; comp_delta_pct = None

    if compare_on:
        comp_from, comp_to = st.date_input("Период сравнения", value=(d_from, d_to),
                                           format="DD.MM.YYYY", key="compare_date_range")
        if comp_from > comp_to: comp_from, comp_to = comp_to, comp_from
        df_comp_base = df.copy()
        for flt_key, flt_col in [("branches","Филиал"),("points","Точки"),
                                  ("categories","Категория"),("subcategories","Подкатегория"),("items","Номенклатура")]:
            vals = st.session_state.applied_filters.get(flt_key, [])
            if vals: df_comp_base = df_comp_base[df_comp_base[flt_col].isin(vals)]
        df_comp = df_comp_base[(df_comp_base["Дата"] >= pd.Timestamp(comp_from)) &
                               (df_comp_base["Дата"] <= pd.Timestamp(comp_to))].copy()
        comp_series = aggregate_for_chart(df_comp, metric_col, pick_freq(comp_from, comp_to))
        comp_total  = float(df_comp[metric_col].sum()) if not df_comp.empty else 0.0
        if comp_total: comp_delta_pct = (cur_total - comp_total) / comp_total * 100.0

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Итого за период", f"{cur_total:,.0f}".replace(",", " "))
    c2.metric("Итого (сравнение)", "—" if not compare_on else f"{comp_total:,.0f}".replace(",", " "))
    c3.metric("Δ к сравнению",    "—" if comp_delta_pct is None else f"{comp_delta_pct:+.1f}%")
    c4.metric("Лучший день", "—" if best_day is None else f"{best_day:%d.%m.%Y} | {best_val:,.0f}".replace(",", " "))

    if cur_series.empty and (not compare_on or comp_series.empty):
        st.info("Нет данных для графика.")
    else:
        fig, ax = plt.subplots(figsize=(12, 4))
        if not cur_series.empty:
            ax.plot(cur_series["Period"], cur_series["Value"], marker="o", linewidth=1, label="Текущий период")
        if compare_on and not comp_series.empty:
            ax.plot(comp_series["Period"], comp_series["Value"], marker="o", linewidth=1, label="Период сравнения")
        ax.set_xlabel("Период"); ax.set_ylabel(metric_col); ax.set_title("Тренд"); ax.legend()
        fig.autofmt_xdate()
        st.pyplot(fig, clear_figure=True)

    st.markdown("**Топ-10 дней (пики) в текущем периоде**")
    if df_daily_cur.empty:
        st.info("Нет данных по дням.")
    else:
        top10 = df_daily_cur.sort_values("Value", ascending=False).head(10).copy()
        top10["Day"] = top10["Day"].dt.date
        st.dataframe(top10[["Day","Value"]], use_container_width=True, hide_index=True)
        trend_sheets = []
        cur_export = cur_series.copy(); cur_export["Period"] = pd.to_datetime(cur_export["Period"])
        trend_sheets.append(("Тренд (текущий)", cur_export, f"Тренд — {d_from:%d.%m.%Y} — {d_to:%d.%m.%Y}"))
        if compare_on and not comp_series.empty:
            comp_export = comp_series.copy(); comp_export["Period"] = pd.to_datetime(comp_export["Period"])
            trend_sheets.append(("Тренд (сравнение)", comp_export, f"Тренд — {comp_from:%d.%m.%Y} — {comp_to:%d.%m.%Y}"))
        top10_exp = df_daily_cur.sort_values("Value", ascending=False).head(10).copy()
        top10_exp["Day"] = pd.to_datetime(top10_exp["Day"])
        trend_sheets.append(("Топ-10 дней", top10_exp, "Топ-10 дней (пики)"))
        download_btn("Скачать тренд и топ-10 дней", trend_sheets,
                     filename=f"trend_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_trend")

# ============================================================
# TAB 5 — KPI по филиалам
# ============================================================
with tab5:
    st.subheader("KPI по филиалам → детализация по точкам")
    if df_filtered.empty:
        st.info("Нет данных по выбранным фильтрам.")
    else:
        st.markdown("### Филиалы (итог)")
        view_b = kpi_branch.copy()
        view_b["Выручка"]      = view_b["Выручка"].round(0)
        view_b["Средний чек"]  = view_b["Средний чек"].round(0)
        view_b["Позиции/чек"]  = view_b["Позиции/чек"].round(2)
        view_b["Товаров/чек"]  = view_b["Товаров/чек"].round(2)
        view_b["Доля выручки"] = (view_b["Доля выручки"] * 100).round(1)
        st.dataframe(view_b, use_container_width=True, hide_index=True)
        st.divider()
        branches_list = sorted(df_filtered["Филиал"].dropna().astype(str).unique().tolist())
        sel_branch    = st.selectbox("Выбери филиал для детализации", branches_list)
        st.markdown(f"### Точки в филиале: {sel_branch}")
        view_p = kpi_branch_point[kpi_branch_point["Филиал"].astype(str) == str(sel_branch)].copy()
        if view_p.empty:
            st.info("В этом филиале нет данных по точкам.")
        else:
            view_p["Выручка"]     = view_p["Выручка"].round(0)
            view_p["Средний чек"] = view_p["Средний чек"].round(0)
            view_p["Позиции/чек"] = view_p["Позиции/чек"].round(2)
            view_p["Товаров/чек"] = view_p["Товаров/чек"].round(2)
            st.dataframe(view_p, use_container_width=True, hide_index=True)
            st.bar_chart(view_p.set_index("Точки")["Выручка"])
        safe_br = sel_branch.replace(" ", "_")[:20]
        download_btn("Скачать KPI по филиалам и точкам",
                     [("KPI Филиалы", kpi_branch,       "KPI по филиалам"),
                      ("KPI Точки",   kpi_branch_point, "KPI по всем точкам"),
                      ("KPI "+sel_branch[:20], view_p,  f"KPI точки — {sel_branch}")],
                     filename=f"kpi_{safe_br}_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_kpi")

# ============================================================
# TAB 6 — Продажи по дням (calendar heatmap)
# ============================================================
with tab6:
    st.subheader("Продажи по дням — календарная карта")
    st.caption("Каждый квадрат — один день. Цвет и цифра показывают объём продаж выбранной позиции.")
    calendar_heatmap_section(df_filtered, metric_col)

# ============================================================
# TAB 7 — Безопасный сток
# ============================================================
with tab7:
    st.subheader("🏭 Неснижаемые остатки — план производства")

    if df_filtered.empty:
        st.info("Нет данных по выбранным фильтрам.")
    else:
        # ── параметры ──────────────────────────────────────────
        st.markdown("#### Параметры")
        col_p1, col_p2, col_p3 = st.columns(3)

        with col_p1:
            coverage_days = st.number_input(
                "Норма остатка (дней)",
                min_value=1, max_value=30, value=9, step=1,
                help="Производственный цикл: 7 дней (до следующего пн) + 2 дня буфер = 9 по умолчанию.",
            )
        with col_p2:
            ss_metric = st.radio(
                "Показывать метрику",
                ["Количество", "Сумма", "Оба"],
                index=0,
                horizontal=True,
            )
        with col_p3:
            abc_filter_ss = st.multiselect(
                "Фильтр по ABC",
                options=["A", "B", "C"],
                default=["A", "B", "C"],
            )

        days_in_period = max((d_to - d_from).days + 1, 1)

        st.caption(
            f"📅 {d_from:%d.%m.%Y} — {d_to:%d.%m.%Y} ({days_in_period} дн.)  "
            f"| 🔄 Производственный план: **каждый понедельник**  "
            f"| 📦 Норма: **{coverage_days} дн.**  "
            f"| Формула: `⌀/день × {coverage_days}`"
        )
        st.divider()

        # ── расчёт ─────────────────────────────────────────────
        ss_df = build_safety_stock(df_filtered, days_in_period, coverage_days)

        if ss_df.empty:
            st.info("Нет данных для расчёта.")
        else:
            # фильтр по ABC
            if abc_filter_ss and "ABC" in ss_df.columns:
                ss_df = ss_df[ss_df["ABC"].isin(abc_filter_ss + ["—"])]

            # выбираем нужные колонки
            base_cols = ["Номенклатура"]
            for c in ["Категория", "Подкатегория"]:
                if c in ss_df.columns:
                    base_cols.append(c)

            if ss_metric == "Количество":
                show_cols = base_cols + ["Итого (кол)", "Среднее/день (кол)",
                                         "Остаток (шт)", "ABC"]
            elif ss_metric == "Сумма":
                show_cols = base_cols + ["Итого (сом)", "Среднее/день (сом)",
                                         "Остаток (сом)", "ABC"]
            else:
                show_cols = base_cols + ["Итого (кол)", "Итого (сом)",
                                         "Среднее/день (кол)", "Среднее/день (сом)",
                                         "Остаток (шт)", "Остаток (сом)", "ABC"]

            show_cols = [c for c in show_cols if c in ss_df.columns]
            display_df = ss_df[show_cols].copy()

            for col in ["Среднее/день (кол)", "Среднее/день (сом)"]:
                if col in display_df.columns:
                    display_df[col] = display_df[col].round(2)
            for col in ["Итого (сом)", "Остаток (сом)"]:
                if col in display_df.columns:
                    display_df[col] = display_df[col].round(0)

            # ── итоги ──────────────────────────────────────────
            m1, m2, m3 = st.columns(3)
            m1.metric("Позиций", len(display_df))
            if "Остаток (шт)" in display_df.columns:
                m2.metric(
                    "Остаток итого (шт)",
                    f"{int(display_df['Остаток (шт)'].sum()):,}".replace(",", " "),
                )
            if "Остаток (сом)" in display_df.columns:
                m3.metric(
                    "Остаток итого (сом)",
                    money(display_df["Остаток (сом)"].sum()),
                )

            # ── таблица ────────────────────────────────────────
            st.markdown("#### Таблица")
            st.dataframe(display_df, use_container_width=True, hide_index=True, height=520)

            # ── экспорт ────────────────────────────────────────
            export_df = ss_df.copy()
            for col in ["Среднее/день (кол)", "Среднее/день (сом)"]:
                if col in export_df.columns:
                    export_df[col] = export_df[col].round(2)
            for col in ["Итого (сом)", "Остаток (сом)"]:
                if col in export_df.columns:
                    export_df[col] = export_df[col].round(0)

            download_btn(
                "Скачать нормы остатков (Excel)",
                [("Нормы остатков", export_df,
                  f"Неснижаемые остатки | {d_from:%d.%m.%Y}–{d_to:%d.%m.%Y} | норма {coverage_days} дн.")],
                filename=f"norms_{d_from:%Y%m%d}_{d_to:%Y%m%d}_cover{coverage_days}d.xlsx",
                key="dl_safety_stock",
            )

            # ── пояснение ──────────────────────────────────────
            with st.expander("ℹ️ Методология расчёта"):
                st.markdown(f"""
**Формула:**
```
Среднее/день  =  Итого за период  ÷  {days_in_period} дней
Неснижаемый остаток  =  ⌈ Среднее/день × {coverage_days} ⌉  (минимум 1 шт)
```

**Почему {coverage_days} дней по умолчанию?**
- Производственный план ставится каждый **понедельник**
- Цикл покрытия = **7 дней** (до следующего понедельника)
- Буфер на время производства и непредвиденный спрос = **+2 дня**
- Итого: **9 дней** покрытия

**Изменить норма** → поле «Норма остатка (дней)» выше.

**Когда добавишь B2B и супермаркеты** — загрузи объединённый файл со всеми каналами,
таблица пересчитается автоматически.
                """)

            # ── РАЗБИВКА НАБОРОВ ───────────────────────────────
            st.divider()
            st.markdown("#### 📦 Разбивка наборов — производство компонентов")

            bom = load_bom()

            if bom.empty:
                st.info(
                    "Файл `разбивка_наборов.xlsx` не найден рядом с `app.py`. "
                    "Положи его рядом — и здесь появится расчёт компонентов и стока наборов."
                )
            else:
                df_comp, df_sets = build_components_breakdown(
                    df_filtered, bom, days_in_period, coverage_days
                )

                col_b1, col_b2 = st.columns(2)

                with col_b1:
                    st.markdown("**Компоненты (производство)**")
                    st.caption(
                        "Прямые продажи + вхождение в наборы. "
                        "Драже — в кг, конфеты — в шт."
                    )
                    if df_comp.empty:
                        st.info("Нет данных о продажах компонентов за период.")
                    else:
                        st.dataframe(df_comp, use_container_width=True,
                                     hide_index=True, height=400)

                with col_b2:
                    st.markdown("**Наборы (сборка)**")
                    st.caption(
                        "Сколько укомплектованных наборов держать на складе. "
                        "Сезонные исключены."
                    )
                    if df_sets.empty:
                        st.info("Нет продаж наборов за выбранный период.")
                    else:
                        st.dataframe(df_sets, use_container_width=True,
                                     hide_index=True, height=400)

                # ── экспорт разбивки ───────────────────────────
                if not df_comp.empty or not df_sets.empty:
                    sheets_breakdown: list[tuple[str, pd.DataFrame, str | None]] = []
                    if not df_comp.empty:
                        sheets_breakdown.append((
                            "Компоненты",
                            df_comp,
                            f"Производство компонентов | {d_from:%d.%m.%Y}–{d_to:%d.%m.%Y} | норма {coverage_days} дн.",
                        ))
                    if not df_sets.empty:
                        sheets_breakdown.append((
                            "Наборы (сборка)",
                            df_sets,
                            f"Сток наборов | {d_from:%d.%m.%Y}–{d_to:%d.%m.%Y} | норма {coverage_days} дн.",
                        ))
                    download_btn(
                        "Скачать компоненты и остатки наборов (Excel)",
                        sheets_breakdown,
                        filename=f"breakdown_{d_from:%Y%m%d}_{d_to:%Y%m%d}_cover{coverage_days}d.xlsx",
                        key="dl_breakdown",
                    )

# ============================================================
# TAB 8 — План / Факт по филиалам
# ============================================================

def _plan_period_block(df_base: pd.DataFrame, block_idx: int, metric_col: str) -> pd.DataFrame | None:
    """
    Рисует один блок: выбор периода + таблица Филиал / Точка / метрики.
    Возвращает итоговый DataFrame для экспорта (или None если данных нет).
    """
    key = f"plan_block_{block_idx}"

    # Дефолтный период — текущий фильтр для блока 0, сдвинутый на неделю для остальных
    default_from = d_from
    default_to   = d_to

    col_date, col_metric = st.columns([3, 1])
    with col_date:
        period = st.date_input(
            "Период",
            value=(default_from, default_to),
            min_value=min_d, max_value=max_d,
            format="DD.MM.YYYY",
            key=f"{key}_dates",
            label_visibility="collapsed",
        )
    with col_metric:
        blk_metric = st.radio(
            "Метрика",
            ["Сумма", "Количество"],
            index=0 if metric_col == "Сумма" else 1,
            horizontal=True,
            key=f"{key}_metric",
            label_visibility="collapsed",
        )

    if not isinstance(period, tuple) or len(period) != 2:
        st.warning("Выбери начало и конец периода.")
        return None

    p_from, p_to = period
    if p_from > p_to:
        p_from, p_to = p_to, p_from

    df_p = df_base[
        (df_base["Дата"] >= pd.Timestamp(p_from)) &
        (df_base["Дата"] <= pd.Timestamp(p_to))
    ].copy()

    if df_p.empty:
        st.info("Нет данных за выбранный период.")
        return None

    blk_col = "Сумма" if blk_metric == "Сумма" else "Количество"
    days_p   = max((p_to - p_from).days + 1, 1)
    label    = "Выручка" if blk_col == "Сумма" else "Кол-во"

    # Агрегат: Филиал → Точка
    g = (df_p.groupby(["Филиал", "Точки"], dropna=False)[blk_col]
         .sum().reset_index()
         .rename(columns={blk_col: label}))
    g["⌀/день"] = (g[label] / days_p).round(1)

    # Итого по филиалу
    branch_total = g.groupby("Филиал")[label].sum().reset_index().rename(columns={label: "Итого филиал"})
    g = g.merge(branch_total, on="Филиал", how="left")
    g["Доля в филиале"] = (g[label] / g["Итого филиал"] * 100).round(1).astype(str) + "%"
    g = g.drop(columns=["Итого филиал"])

    # Сортировка: по филиалу, внутри — по убыванию
    g = g.sort_values(["Филиал", label], ascending=[True, False]).reset_index(drop=True)

    # Итоговая строка
    total_row = pd.DataFrame([{
        "Филиал": "ИТОГО",
        "Точки": "—",
        label: g[label].sum(),
        "⌀/день": round(g[label].sum() / days_p, 1),
        "Доля в филиале": "100%",
    }])
    g_display = pd.concat([g, total_row], ignore_index=True)

    # Форматирование числа
    if blk_col == "Сумма":
        g_display[label]    = g_display[label].apply(lambda x: f"{x:,.0f}".replace(",", " "))
        g_display["⌀/день"] = g_display["⌀/день"].apply(lambda x: f"{x:,.0f}".replace(",", " "))
    else:
        g_display[label]    = g_display[label].apply(lambda x: f"{int(x):,}".replace(",", " "))
        g_display["⌀/день"] = g_display["⌀/день"].apply(lambda x: f"{x:.1f}")

    st.caption(f"📅 {p_from:%d.%m.%Y} — {p_to:%d.%m.%Y}  |  {days_p} дн.  |  {label}")
    st.dataframe(g_display, use_container_width=True, hide_index=True, height=min(35 * len(g_display) + 38, 520))

    # Возвращаем чистый df для экспорта
    export = g.copy()
    export.insert(0, "Период", f"{p_from:%d.%m.%Y}–{p_to:%d.%m.%Y}")
    return export


with tab8:
    st.subheader("📋 Факт по филиалам — сравнение периодов")
    st.caption(
        "Каждый блок — отдельный период. "
        "Добавляй блоки чтобы сравнивать факт за разные недели/месяцы и ставить план."
    )

    if df_filtered.empty:
        st.info("Нет данных по выбранным фильтрам.")
    else:
        # Используем df без фильтра по дате — дата выбирается внутри блока
        df_plan_base = df.copy()
        df_plan_base = basic_clean(df_plan_base)

        # Применяем только НЕ-датовые фильтры из сайдбара
        if branches_selected:
            df_plan_base = df_plan_base[df_plan_base["Филиал"].isin(branches_selected)]
        if points_selected:
            df_plan_base = df_plan_base[df_plan_base["Точки"].isin(points_selected)]
        if categories_selected:
            df_plan_base = df_plan_base[df_plan_base["Категория"].isin(categories_selected)]
        if subcategories_selected:
            df_plan_base = df_plan_base[df_plan_base["Подкатегория"].isin(subcategories_selected)]
        if items_selected:
            df_plan_base = df_plan_base[df_plan_base["Номенклатура"].isin(items_selected)]

        # Счётчик блоков
        if "plan_block_count" not in st.session_state:
            st.session_state["plan_block_count"] = 2

        n_blocks = st.session_state["plan_block_count"]

        # Кнопки управления сверху
        btn1, btn2, btn3 = st.columns([1, 1, 6])
        with btn1:
            if st.button("＋ Добавить период", key="plan_add", use_container_width=True):
                st.session_state["plan_block_count"] += 1
                st.rerun()
        with btn2:
            if n_blocks > 1:
                if st.button("－ Убрать", key="plan_remove", use_container_width=True):
                    st.session_state["plan_block_count"] -= 1
                    st.rerun()

        st.divider()

        # Рисуем блоки в колонках по 2
        export_frames = []
        pairs = [(i, i+1) for i in range(0, n_blocks, 2)]

        for pair in pairs:
            cols = st.columns(len([p for p in pair if p < n_blocks]), gap="large")
            for col_idx, block_idx in enumerate(pair):
                if block_idx >= n_blocks:
                    break
                with cols[col_idx]:
                    # Заголовок блока
                    st.markdown(
                        f"<div style='background:#1F4E79;border-radius:6px;"
                        f"padding:4px 12px;margin-bottom:8px;'>"
                        f"<span style='color:white;font-weight:700;font-size:13px;'>"
                        f"Период {block_idx + 1}</span></div>",
                        unsafe_allow_html=True,
                    )
                    result = _plan_period_block(df_plan_base, block_idx, metric_col)
                    if result is not None:
                        export_frames.append((f"Период {block_idx+1}", result))

            if pair != pairs[-1]:
                st.divider()

        # Экспорт всех периодов одним файлом
        if export_frames:
            st.divider()
            sheets_plan = [
                (name, frame, f"Факт по точкам — {name}")
                for name, frame in export_frames
            ]
            # Добавляем сводный лист — все периоды вертикально
            all_combined = pd.concat([f for _, f in export_frames], ignore_index=True)
            sheets_plan.append(("Все периоды", all_combined, "Сводная — все периоды"))

            download_btn(
                "Скачать факт по всем периодам (Excel)",
                sheets_plan,
                filename=f"plan_fact_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx",
                key="dl_plan_fact",
            )