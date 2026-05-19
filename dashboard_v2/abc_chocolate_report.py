"""ABC-анализ шоколадной продукции 2025 vs 2026 (по месяцам) + себес и маржа.

Источники:
  - docs/база.parquet        — продажи
  - D:/claude/себес шоколад.xlsx — справочник себестоимости и розничной цены

На выходе: D:/dashboard/reports/ABC_chocolate_2025_2026.xlsx
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

ROOT = Path(__file__).resolve().parents[1]
SALES_PATH = ROOT / "docs" / "база.parquet"
COST_PATH = Path("D:/claude/себес шоколад.xlsx")
OUT_PATH = ROOT / "reports" / "ABC_chocolate_2025_2026.xlsx"

MONTHS_RU = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
             "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]

# ---------- нормализация имён ----------

_PAREN_1SHT = re.compile(r"\(\s*1\s*(?:шт|конфета)\s*\)", re.IGNORECASE)
_SPACES = re.compile(r"\s+")


def norm(s: object) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    t = str(s).lower().strip()
    t = t.replace("ё", "е").replace('"', "").replace("'", "")
    t = _PAREN_1SHT.sub(" ", t)
    t = t.replace(".", " ").replace(",", " ")
    t = _SPACES.sub(" ", t).strip()
    return t


def variants(s: str) -> list[str]:
    base = norm(s)
    out = [base]
    if base.startswith("плитка "):
        out.append(base[len("плитка "):])
    else:
        out.append("плитка " + base)
    return out


# ---------- загрузка ----------

# Шоколадные наборы из «Наборы и коробки» — биз-логически это шоколад,
# поэтому включаем в ABC шоколада.
CHOCOLATE_SET_PREFIXES = (
    "Нац.коллекция",
    "Корпусная коллекция",
    "Нарезная колллекция",
    "Коктельная коллекция",
    "Набор из дубайских",
    "Слива-Грецкий орех корпусная НК",
    "Ынтымак Манас Ордо",
)

# Последняя дата с полными данными (18.05.2026 — есть только у одного филиала, пропускаем).
DATA_CUTOFF = pd.Timestamp("2026-05-17")


def is_chocolate(row_group: str, row_name: str) -> bool:
    if row_group == "Шоколад и конфеты":
        return True
    if row_group == "Наборы и коробки" and str(row_name).startswith(CHOCOLATE_SET_PREFIXES):
        return True
    return False


def load_sales() -> pd.DataFrame:
    df = pd.read_parquet(SALES_PATH)
    df["Дата_p"] = pd.to_datetime(df["Дата"], format="%d.%m.%Y", errors="coerce")
    df = df.dropna(subset=["Дата_p"])
    # Обрезаем неполный день 18.05.2026 — данные есть только у одного филиала.
    df = df[df["Дата_p"] <= DATA_CUTOFF].copy()

    # Скоуп шоколада: основная группа + биз-логически шоколадные наборы.
    name_starts = df["Номенклатура"].astype(str).str.startswith(CHOCOLATE_SET_PREFIXES)
    keep = (df["Группа"] == "Шоколад и конфеты") | (
        (df["Группа"] == "Наборы и коробки") & name_starts
    )
    df = df[keep].copy()

    df["Год"] = df["Дата_p"].dt.year
    df["Мес"] = df["Дата_p"].dt.month
    df["Количество"] = pd.to_numeric(df["Количество"], errors="coerce").fillna(0)
    df["Сумма"] = pd.to_numeric(df["Сумма"], errors="coerce").fillna(0)
    return df


def load_cost() -> pd.DataFrame:
    c = pd.read_excel(COST_PATH)
    c = c.rename(columns={c.columns[0]: "Номенклатура"})
    c = c.dropna(subset=["Номенклатура"]).copy()
    for col in ["себес сырьевая", "пнр", "себес+пнр", "розница", "маржа сом", "%маржа"]:
        if col in c.columns:
            c[col] = pd.to_numeric(c[col], errors="coerce")
    c["key"] = c["Номенклатура"].map(norm)
    return c


def build_cost_lookup(cost: pd.DataFrame) -> dict[str, dict]:
    """Возвращает dict: ключ-нормализованное-имя → запись себеса."""
    lookup: dict[str, dict] = {}
    for _, row in cost.iterrows():
        rec = {
            "cost_name": row["Номенклатура"],
            "себес_сырье": row.get("себес сырьевая", np.nan),
            "пнр": row.get("пнр", np.nan),
            "себес_полн": row.get("себес+пнр", np.nan),
            "розница_прайс": row.get("розница", np.nan),
            "маржа_прайс": row.get("маржа сом", np.nan),
            "%маржа_прайс": row.get("%маржа", np.nan),
        }
        for k in variants(row["Номенклатура"]):
            if k and k not in lookup:
                lookup[k] = rec
    return lookup


def attach_cost(skus: list[str], lookup: dict[str, dict]) -> pd.DataFrame:
    rows = []
    for sku in skus:
        rec = None
        for k in variants(sku):
            if k in lookup:
                rec = lookup[k]
                break
        rows.append({
            "Номенклатура": sku,
            "Себес сырье": rec["себес_сырье"] if rec else np.nan,
            "ПНР": rec["пнр"] if rec else np.nan,
            "Себес полн.": rec["себес_полн"] if rec else np.nan,
            "Прайс розница": rec["розница_прайс"] if rec else np.nan,
            "Совпало с прайсом": "да" if rec else "нет",
        })
    return pd.DataFrame(rows)


# ---------- ABC ----------

def abc_label(cum_share: float) -> str:
    if cum_share <= 0.80:
        return "A"
    if cum_share <= 0.95:
        return "B"
    return "C"


def build_period(df: pd.DataFrame, cost_df: pd.DataFrame, label: str) -> pd.DataFrame:
    """Агрегат по SKU за период + ABC по выручке."""
    if df.empty:
        return pd.DataFrame()
    g = (df.groupby("Номенклатура", dropna=False)
           .agg(Количество=("Количество", "sum"),
                Выручка=("Сумма", "sum"))
           .reset_index())
    g["Ср. цена"] = np.where(g["Количество"] > 0,
                              g["Выручка"] / g["Количество"], np.nan)
    g = g.sort_values("Выручка", ascending=False).reset_index(drop=True)
    total = g["Выручка"].sum()
    g["Доля выручки"] = np.where(total > 0, g["Выручка"] / total, 0.0)
    g["Кум. доля"] = g["Доля выручки"].cumsum()
    g["ABC"] = g["Кум. доля"].map(abc_label)
    g = g.merge(cost_df, on="Номенклатура", how="left")
    g["Маржа за шт"] = g["Ср. цена"] - g["Себес полн."]
    g["Маржа сумма"] = g["Маржа за шт"] * g["Количество"]
    g["% маржи"] = np.where(g["Ср. цена"].fillna(0) > 0,
                             g["Маржа за шт"] / g["Ср. цена"], np.nan)
    g.insert(0, "Период", label)
    cols = ["Период", "Номенклатура", "Количество", "Выручка",
            "Доля выручки", "Кум. доля", "ABC",
            "Ср. цена", "Себес сырье", "ПНР", "Себес полн.",
            "Маржа за шт", "% маржи", "Маржа сумма",
            "Прайс розница", "Совпало с прайсом"]
    return g[cols]


def build_year_monthly(df_year: pd.DataFrame, cost_df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Длинный df: строка = (SKU, месяц) с ABC внутри месяца."""
    parts = []
    for m in sorted(df_year["Мес"].unique()):
        sub = df_year[df_year["Мес"] == m]
        if sub.empty:
            continue
        part = build_period(sub, cost_df, f"{MONTHS_RU[m-1]} {year}")
        part.insert(1, "Месяц №", m)
        parts.append(part)
    if not parts:
        return pd.DataFrame()
    return pd.concat(parts, ignore_index=True)


def build_year_summary(df_year: pd.DataFrame, cost_df: pd.DataFrame, year: int) -> pd.DataFrame:
    return build_period(df_year, cost_df, f"Итого {year}")


# ---------- сравнение 25 vs 26 ----------

def build_compare(monthly_25: pd.DataFrame,
                  monthly_26: pd.DataFrame,
                  cost_df: pd.DataFrame) -> pd.DataFrame:
    if monthly_25.empty or monthly_26.empty:
        return pd.DataFrame()
    months_in_26 = sorted(monthly_26["Месяц №"].unique())
    a = (monthly_25[monthly_25["Месяц №"].isin(months_in_26)]
         .groupby(["Месяц №", "Номенклатура"], as_index=False)
         .agg(Кол_25=("Количество", "sum"),
              Выр_25=("Выручка", "sum"),
              ABC_25=("ABC", "first"),
              Маржа_25=("Маржа сумма", "sum")))
    b = (monthly_26.groupby(["Месяц №", "Номенклатура"], as_index=False)
         .agg(Кол_26=("Количество", "sum"),
              Выр_26=("Выручка", "sum"),
              ABC_26=("ABC", "first"),
              Маржа_26=("Маржа сумма", "sum")))
    m = a.merge(b, on=["Месяц №", "Номенклатура"], how="outer")
    for c in ["Кол_25", "Кол_26", "Выр_25", "Выр_26", "Маржа_25", "Маржа_26"]:
        m[c] = m[c].fillna(0)
    m["Δ кол."] = m["Кол_26"] - m["Кол_25"]
    m["Δ выр. сом"] = m["Выр_26"] - m["Выр_25"]
    m["Δ выр. %"] = np.where(m["Выр_25"] > 0,
                               m["Δ выр. сом"] / m["Выр_25"], np.nan)
    m["Δ маржа сом"] = m["Маржа_26"] - m["Маржа_25"]
    m["Месяц"] = m["Месяц №"].map(lambda x: MONTHS_RU[int(x)-1])
    m = m.merge(cost_df[["Номенклатура", "Себес полн.", "Совпало с прайсом"]],
                on="Номенклатура", how="left")
    cols = ["Месяц №", "Месяц", "Номенклатура",
            "Кол_25", "Кол_26", "Δ кол.",
            "Выр_25", "Выр_26", "Δ выр. сом", "Δ выр. %",
            "ABC_25", "ABC_26",
            "Маржа_25", "Маржа_26", "Δ маржа сом",
            "Себес полн.", "Совпало с прайсом"]
    return m[cols].sort_values(["Месяц №", "Выр_26"], ascending=[True, False])


# ---------- кандидаты на вывод ----------

def build_candidates(s25_same: pd.DataFrame,
                     s26: pd.DataFrame,
                     cost_df: pd.DataFrame) -> pd.DataFrame:
    """SKU-кандидаты на снятие.

    Сравнение сопоставимое: 2025 vs 2026 только по месяцам, что есть в 2026.

    Сигналы:
      * ABC = C в 2026 YTD
      * Маржа за шт <= 0  (если есть себес)
      * % маржи < 0.30   (если есть себес)
      * Падение выручки 26 vs 25 (same period) более 30%
    """
    if s26.empty:
        return pd.DataFrame()
    sum_25 = build_period(s25_same, cost_df, "2025 (сопост.)") if not s25_same.empty else pd.DataFrame()
    sum_26 = build_period(s26, cost_df, "2026 YTD")
    df = sum_26[["Номенклатура", "Количество", "Выручка", "ABC",
                  "Ср. цена", "Себес полн.", "Маржа за шт",
                  "% маржи", "Маржа сумма", "Совпало с прайсом"]].copy()
    df = df.rename(columns={"Количество": "Кол. 26 YTD",
                              "Выручка": "Выр. 26 YTD",
                              "ABC": "ABC 26",
                              "Ср. цена": "Ср. цена 26",
                              "Маржа сумма": "Маржа 26 YTD"})
    if not sum_25.empty:
        df = df.merge(sum_25[["Номенклатура", "Выручка", "ABC"]]
                        .rename(columns={"Выручка": "Выр. 25 (сопост.)",
                                          "ABC": "ABC 25 (сопост.)"}),
                        on="Номенклатура", how="left")
    else:
        df["Выр. 25 (сопост.)"] = np.nan
        df["ABC 25 (сопост.)"] = np.nan
    df["Падение выр. %"] = np.where(df["Выр. 25 (сопост.)"].fillna(0) > 0,
                                      (df["Выр. 26 YTD"] - df["Выр. 25 (сопост.)"]) / df["Выр. 25 (сопост.)"],
                                      np.nan)
    sig = []
    flags = []
    for _, r in df.iterrows():
        f = []
        if r["ABC 26"] == "C":
            f.append("ABC=C")
        if pd.notna(r["Маржа за шт"]) and r["Маржа за шт"] <= 0:
            f.append("маржа≤0")
        if pd.notna(r["% маржи"]) and r["% маржи"] < 0.30:
            f.append("%маржи<30")
        if pd.notna(r["Падение выр. %"]) and r["Падение выр. %"] < -0.30:
            f.append("выр.−30% (сопост.)")
        flags.append(", ".join(f))
        sig.append(len(f))
    df["Сигналов"] = sig
    df["Флаги"] = flags
    cand = df[df["Сигналов"] >= 2].copy()
    cand = cand.sort_values(["Сигналов", "Выр. 26 YTD"], ascending=[False, True])
    cols = ["Номенклатура", "Сигналов", "Флаги",
            "ABC 25 (сопост.)", "ABC 26",
            "Кол. 26 YTD", "Выр. 25 (сопост.)", "Выр. 26 YTD", "Падение выр. %",
            "Ср. цена 26", "Себес полн.", "Маржа за шт", "% маржи",
            "Маржа 26 YTD", "Совпало с прайсом"]
    return cand[cols]


# ---------- запись Excel ----------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ABC_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="FFEB9C"),
    "C": PatternFill("solid", fgColor="FFC7CE"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)

MONEY_COLS = {"Выручка", "Маржа сумма", "Маржа за шт", "Себес сырье", "ПНР",
              "Себес полн.", "Ср. цена", "Прайс розница",
              "Выр_25", "Выр_26", "Δ выр. сом", "Маржа_25", "Маржа_26", "Δ маржа сом",
              "Выр. 25 (сопост.)", "Выр. 26 YTD", "Маржа 26 YTD", "Ср. цена 26"}
INT_COLS = {"Количество", "Кол_25", "Кол_26", "Δ кол.", "Месяц №",
            "Кол. 26 YTD", "Сигналов"}
PCT_COLS = {"Доля выручки", "Кум. доля", "% маржи", "Δ выр. %", "Падение выр. %"}


def write_sheet(ws, df: pd.DataFrame, abc_col: str | None = None) -> None:
    if df.empty:
        ws["A1"] = "Нет данных"
        return
    headers = list(df.columns)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=str(h))
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    abc_idx = headers.index(abc_col) + 1 if abc_col and abc_col in headers else None

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, (col, v) in enumerate(zip(headers, row), 1):
            cell = ws.cell(row=ri, column=ci)
            if v is None or (isinstance(v, float) and pd.isna(v)) or v is pd.NA:
                cell.value = None
            elif hasattr(v, "item"):
                try:
                    cell.value = v.item()
                except (ValueError, AttributeError):
                    cell.value = str(v)
            else:
                cell.value = v
            if col in MONEY_COLS:
                cell.number_format = '#,##0'
            elif col in INT_COLS:
                cell.number_format = '#,##0'
            elif col in PCT_COLS:
                cell.number_format = '0.0%'
            cell.font = Font(size=10)
            cell.border = BORDER
        if abc_idx is not None:
            v = ws.cell(row=ri, column=abc_idx).value
            if v in ABC_FILL:
                ws.cell(row=ri, column=abc_idx).fill = ABC_FILL[v]

    # autowidth
    for col in ws.columns:
        ml = 0
        letter = get_column_letter(col[0].column)
        for c in col:
            txt = "" if c.value is None else str(c.value)
            if len(txt) > ml:
                ml = len(txt)
        ws.column_dimensions[letter].width = min(max(ml + 2, 8), 48)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------- main ----------

def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    print("Загружаю продажи…")
    sales = load_sales()
    print(f"  строк: {len(sales):,}, SKU: {sales['Номенклатура'].nunique()}")

    print("Загружаю себес…")
    cost_raw = load_cost()
    lookup = build_cost_lookup(cost_raw)

    all_skus = sales["Номенклатура"].unique().tolist()
    cost_df = attach_cost(all_skus, lookup)
    matched = (cost_df["Совпало с прайсом"] == "да").sum()
    print(f"  покрытие себесом: {matched}/{len(cost_df)} SKU")

    s25 = sales[sales["Год"] == 2025].copy()
    s26 = sales[sales["Год"] == 2026].copy()
    print(f"2025: {len(s25):,} строк, месяцев {s25['Мес'].nunique()}")
    print(f"2026: {len(s26):,} строк, месяцев {s26['Мес'].nunique()}")

    # --- сопоставимый период: 2025 усекаем под фактический диапазон 2026 ---
    # Последний месяц 2026 может быть неполным (например, 1–18 мая).
    # Берём максимальную дату 2026, и для 2025 в этот же месяц оставляем дни 1..max_day.
    max_date_26 = s26["Дата_p"].max()
    last_month_26 = int(max_date_26.month)
    last_day_26 = int(max_date_26.day)
    full_months_26 = sorted(m for m in s26["Мес"].unique() if m != last_month_26)
    s25_full_months = s25[s25["Мес"].isin(full_months_26)]
    s25_partial = s25[(s25["Мес"] == last_month_26) & (s25["Дата_p"].dt.day <= last_day_26)]
    s25_same = pd.concat([s25_full_months, s25_partial], ignore_index=True)
    print(f"Сопоставимый период 2025: до {last_day_26:02d}.{last_month_26:02d}.2025 "
          f"({len(s25_same):,} строк)")

    monthly_25 = build_year_monthly(s25, cost_df, 2025)            # полный 2025 — для листа «по месяцам»
    monthly_25_same = build_year_monthly(s25_same, cost_df, 2025)  # усечённый — для сравнения
    monthly_26 = build_year_monthly(s26, cost_df, 2026)
    sum_25 = build_year_summary(s25, cost_df, 2025)
    sum_26 = build_year_summary(s26, cost_df, 2026)
    compare = build_compare(monthly_25_same, monthly_26, cost_df)
    sum_25_same = build_period(s25_same, cost_df, "2025 (сопост. с 26)") if not s25_same.empty else pd.DataFrame()
    cand = build_candidates(s25_same, s26, cost_df)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws = wb.create_sheet("Summary")
    ws["A1"] = "ABC-анализ шоколадной продукции — 2025 vs 2026"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A3"] = f"SKU в продажах:  {len(all_skus)}"
    ws["A4"] = f"Совпало с прайсом себеса:  {matched} (см. колонку «Совпало с прайсом»)"
    ws["A5"] = f"2025: месяцев {s25['Мес'].nunique()}, выручка {int(s25['Сумма'].sum()):,} сом"
    ws["A6"] = (f"2026: до {last_day_26:02d}.{last_month_26:02d}.2026 "
                f"({s26['Мес'].nunique()} мес., последний — неполный), "
                f"выручка {int(s26['Сумма'].sum()):,} сом")
    ws["A7"] = (f"Сопоставимый период 2025: 01.{full_months_26[0]:02d}.2025 – "
                f"{last_day_26:02d}.{last_month_26:02d}.2025  "
                f"(для листов «ABC 2025 сопост.», «Сравнение 25 vs 26», «Кандидаты»).")
    ws["A9"] = "Логика ABC по выручке внутри периода: A ≤ 80%, B ≤ 95%, C ≤ 100% кумулятивной доли."
    ws["A10"] = "Маржа = Ср.цена − Себес полн.  (где есть совпадение по прайсу)."
    ws["A11"] = "Кандидаты на вывод: ≥ 2 сигналов (ABC=C, маржа≤0, %маржи<30, падение выручки >30%)."
    ws.column_dimensions["A"].width = 100

    write_sheet(wb.create_sheet("ABC 2025 (год)"), sum_25, abc_col="ABC")
    write_sheet(wb.create_sheet("ABC 2025 сопост."), sum_25_same, abc_col="ABC")
    write_sheet(wb.create_sheet("ABC 2026 (YTD)"), sum_26, abc_col="ABC")
    write_sheet(wb.create_sheet("ABC 2025 по месяцам"), monthly_25, abc_col="ABC")
    write_sheet(wb.create_sheet("ABC 2026 по месяцам"), monthly_26, abc_col="ABC")
    write_sheet(wb.create_sheet("Сравнение 25 vs 26"), compare, abc_col="ABC_26")
    write_sheet(wb.create_sheet("Кандидаты на вывод"), cand, abc_col="ABC 26")

    wb.save(OUT_PATH)
    print(f"\nГотово: {OUT_PATH}")
    print(f"Файл: {OUT_PATH.stat().st_size/1024:.1f} КБ")


if __name__ == "__main__":
    main()
