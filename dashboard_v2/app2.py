"""
Sales Dashboard v2 — умные решения для розницы.

Новое относительно v1:
  - Главная с автоинсайтами (топ-роста, топ-падения, мёртвые SKU, аномалии)
  - Сравнение периодов MoM / YoY / WoW автоматом
  - Cross-sell (ассоциации товаров внутри чека)
  - Прогноз спроса 7/14/30 дней (сезонный скользящий)
  - Поиск SKU + карточка товара
  - Рейтинг филиалов и точек по всем KPI
  - Календарь сезонности категорий (месяц × группа)
  - Группировка категорий в 8 верхнеуровневых групп

Источник данных: GitHub Pages (этот же репо), fallback — локальный файл.
"""
from __future__ import annotations

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
import matplotlib.dates as mdates
import io
import requests
import calendar
from pathlib import Path
from datetime import date, timedelta
from urllib.parse import quote
from itertools import combinations
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# импорт маппинга из соседнего файла
import sys
sys.path.insert(0, str(Path(__file__).parent))
from category_mapping import (
    apply_mapping, category_to_group, normalize_subcategory,
    GROUP_ORDER, GROUP_SEASONAL,
)

# =============================================================================
# Page config + стили
# =============================================================================
st.set_page_config(
    page_title="Sales Dashboard v2",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  /* компактнее заголовки вкладок */
  .stTabs [data-baseweb="tab-list"] { gap: 2px; flex-wrap: wrap; }
  .stTabs [data-baseweb="tab"] { padding: 6px 12px; font-size: 14px; }
  /* карточки KPI */
  div[data-testid="stMetric"] {
      background: #FAFBFC;
      border: 1px solid #E1E4E8;
      border-radius: 8px;
      padding: 10px 14px;
  }
  /* инсайт-карточки */
  .insight-card {
      background: linear-gradient(135deg, #fafbfd 0%, #f0f4fa 100%);
      border-left: 4px solid #1F4E79;
      border-radius: 6px;
      padding: 10px 14px;
      margin-bottom: 10px;
  }
  .insight-card.warn   { border-left-color: #E67E22; }
  .insight-card.danger { border-left-color: #C0392B; }
  .insight-card.ok     { border-left-color: #27AE60; }
  .insight-title { font-weight: 700; font-size: 13px; color: #1F3864; margin-bottom: 4px; }
  .insight-body  { font-size: 12px; color: #333; line-height: 1.5; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# Источник данных
# =============================================================================
GH_USER = "MonkeyDjack"
GH_REPO = "yahya-sales-dashboard"
EXCEL_NAME = "Итоговый_отчет1.xlsx"
BOM_NAME   = "разбивка_наборов.xlsx"

GH_PAGES_EXCEL = f"https://{GH_USER.lower()}.github.io/{GH_REPO}/{quote(EXCEL_NAME)}"
GH_PAGES_BOM   = f"https://{GH_USER.lower()}.github.io/{GH_REPO}/{quote(BOM_NAME)}"
GDRIVE_FILE_ID = "1FLoz9fyHlAke0MgrEgwSd8eTekH-zbCc"


@st.cache_data(ttl=3600, show_spinner="Загрузка с GitHub Pages...")
def load_from_github_pages(url: str) -> bytes:
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content


@st.cache_data(ttl=3600, show_spinner="Загрузка с Google Sheets...")
def load_from_gdrive(file_id: str) -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content


def read_main_sheet(xlsx_bytes: bytes) -> pd.DataFrame:
    bio = io.BytesIO(xlsx_bytes)
    xls = pd.ExcelFile(bio)
    preferred = ["база", "База", "Sheet1", "Лист1", "Лист 1"]
    sheet = next((s for s in preferred if s in xls.sheet_names), xls.sheet_names[0])
    df = pd.read_excel(bio, sheet_name=sheet, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def basic_clean(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Дата" in df.columns:
        dt = pd.to_datetime(df["Дата"], errors="coerce", dayfirst=True)
        df["Дата"] = dt.dt.normalize()
        df = df[df["Дата"].notna()]
    for col in ["Количество", "Сумма", "Цена"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in ["Филиал", "Точки", "Номенклатура", "Категория",
                "Подкатегория", "Группа", "Время"]:
        if col in df.columns:
            df[col] = df[col].astype("string").str.strip()
    # страховка: если Группа отсутствует или пустая — досчитаем в памяти
    df = apply_mapping(df)
    return df


def validate_minimum(df: pd.DataFrame) -> None:
    required = ["Филиал", "Точки", "Номенклатура", "Количество", "Сумма", "Дата"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"В данных не хватает колонок: {missing}")
        st.stop()
    if df["Дата"].isna().all():
        st.error("Не удалось распознать ни одной даты.")
        st.stop()


# =============================================================================
# Sidebar: источник
# =============================================================================
st.sidebar.header("Источник данных")
source_mode = st.sidebar.radio(
    "Откуда грузить?",
    ["GitHub Pages", "Google Drive", "Локальный файл", "Загрузить вручную"],
    index=0,
    help="GitHub Pages — это репо с обновляемым Excel. "
         "Google Drive — резерв. Локальный — когда разработчик сам запускает.",
)

df = None
if source_mode == "GitHub Pages":
    try:
        xlsx_bytes = load_from_github_pages(GH_PAGES_EXCEL)
        df = read_main_sheet(xlsx_bytes)
        st.sidebar.success(f"✓ GitHub Pages ({len(xlsx_bytes)/1024/1024:.1f} МБ)")
    except Exception as e:
        st.error(f"❌ Не удалось загрузить с GitHub Pages ({GH_PAGES_EXCEL}).\n\n{e}")
        st.info("Проверь что Pages включён в Settings → Pages (Source: main, /docs).")
        st.stop()

elif source_mode == "Google Drive":
    try:
        xlsx_bytes = load_from_gdrive(GDRIVE_FILE_ID)
        df = read_main_sheet(xlsx_bytes)
        st.sidebar.success(f"✓ Google Drive")
    except Exception as e:
        st.error(f"Ошибка Google Drive: {e}")
        st.stop()

elif source_mode == "Локальный файл":
    local_candidates = [
        Path(__file__).parent / "docs" / EXCEL_NAME,
        Path(__file__).parent.parent / EXCEL_NAME,
        Path(EXCEL_NAME),
    ]
    path = next((p for p in local_candidates if p.exists()), None)
    if path is None:
        st.error("Локальный Excel не найден.")
        st.stop()
    df = pd.read_excel(str(path), sheet_name="база", engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    st.sidebar.success(f"✓ {path}")

else:
    up = st.sidebar.file_uploader("Excel (.xlsx)", type=["xlsx"])
    if up is not None:
        st.session_state["uploaded_xlsx"] = up.getvalue()
    if "uploaded_xlsx" in st.session_state:
        df = read_main_sheet(st.session_state["uploaded_xlsx"])

if df is None:
    st.info("Выбери источник данных.")
    st.stop()

df = basic_clean(df)
validate_minimum(df)

# =============================================================================
# Фильтры (форма, чтобы не пересчитывать на каждый клик)
# =============================================================================
min_ts = df["Дата"].min()
max_ts = df["Дата"].max()
min_d  = min_ts.date()
max_d  = max_ts.date()

branches_all = sorted(df["Филиал"].dropna().astype(str).unique().tolist())
groups_all   = [g for g in GROUP_ORDER if g in df["Группа"].dropna().astype(str).unique()]

if "ap" not in st.session_state:
    st.session_state.ap = {
        "date_range":     (min_d, max_d),
        "branches":       branches_all,
        "points":         [],
        "groups":         [],
        "categories":     [],
        "subcategories":  [],
        "items":          [],
        "abc_metric":     "Сумма",
    }

if "fv" not in st.session_state:
    st.session_state.fv = 0   # filters version — чтобы форсить reset виджетов

# Санация: если границы дат / филиалы изменились — чиним
ap = st.session_state.ap
if (ap["date_range"][0] < min_d) or (ap["date_range"][1] > max_d) \
        or not set(ap["branches"]).issubset(set(branches_all)):
    ap["date_range"] = (min_d, max_d)
    ap["branches"] = branches_all
    ap["points"] = []
    st.session_state.fv += 1

# вспомогательный срез для зависимых мультиселектов
def _cascade_options(df_src, drafts: dict) -> dict:
    out = {}
    base = df_src.copy()
    if drafts.get("branches"):
        base = base[base["Филиал"].isin(drafts["branches"])]
    out["points"] = sorted(base["Точки"].dropna().astype(str).unique().tolist())
    if drafts.get("points"):
        base = base[base["Точки"].isin(drafts["points"])]
    out["groups"] = [g for g in GROUP_ORDER if g in base["Группа"].dropna().unique()]
    if drafts.get("groups"):
        base = base[base["Группа"].isin(drafts["groups"])]
    out["categories"] = sorted(base["Категория"].dropna().astype(str).unique().tolist())
    if drafts.get("categories"):
        base = base[base["Категория"].isin(drafts["categories"])]
    out["subcategories"] = sorted(base["Подкатегория"].dropna().astype(str).unique().tolist())
    if drafts.get("subcategories"):
        base = base[base["Подкатегория"].isin(drafts["subcategories"])]
    out["items"] = sorted(base["Номенклатура"].dropna().astype(str).unique().tolist())
    return out


st.sidebar.header("Фильтры")
with st.sidebar.form("filters_form", clear_on_submit=False):
    fv = st.session_state.fv
    d_def = ap["date_range"]
    draft_dates = st.date_input(
        "Период",
        value=d_def, min_value=min_d, max_value=max_d,
        format="DD.MM.YYYY", key=f"dt_{fv}",
    )
    draft_branches = st.multiselect(
        "Филиал", branches_all, default=ap["branches"], key=f"br_{fv}",
    )

    _opts = _cascade_options(df, {"branches": draft_branches})
    draft_points = st.multiselect(
        "Точки", _opts["points"],
        default=[p for p in ap["points"] if p in _opts["points"]],
        key=f"pt_{fv}",
    )

    _opts = _cascade_options(df, {"branches": draft_branches, "points": draft_points})
    draft_groups = st.multiselect(
        "Группа (верхний уровень)", _opts["groups"],
        default=[g for g in ap["groups"] if g in _opts["groups"]],
        key=f"gp_{fv}",
    )

    _opts = _cascade_options(df, {
        "branches": draft_branches, "points": draft_points,
        "groups": draft_groups,
    })
    draft_cats = st.multiselect(
        "Категория", _opts["categories"],
        default=[c for c in ap["categories"] if c in _opts["categories"]],
        key=f"cat_{fv}",
    )

    _opts = _cascade_options(df, {
        "branches": draft_branches, "points": draft_points,
        "groups": draft_groups, "categories": draft_cats,
    })
    draft_subs = st.multiselect(
        "Подкатегория", _opts["subcategories"],
        default=[s for s in ap["subcategories"] if s in _opts["subcategories"]],
        key=f"sub_{fv}",
    )

    _opts = _cascade_options(df, {
        "branches": draft_branches, "points": draft_points,
        "groups": draft_groups, "categories": draft_cats,
        "subcategories": draft_subs,
    })
    draft_items = st.multiselect(
        "Номенклатура", _opts["items"],
        default=[i for i in ap["items"] if i in _opts["items"]],
        key=f"it_{fv}", help="Пусто — все SKU.",
    )

    draft_metric = st.radio(
        "Метрика ABC", ["Сумма", "Количество"],
        index=0 if ap["abc_metric"] == "Сумма" else 1,
        horizontal=True, key=f"mt_{fv}",
    )

    if st.form_submit_button("Применить", use_container_width=True, type="primary"):
        if isinstance(draft_dates, tuple) and len(draft_dates) == 2:
            d_from, d_to = draft_dates
        else:
            d_from = d_to = draft_dates
        if d_from > d_to: d_from, d_to = d_to, d_from
        if not draft_branches: draft_branches = branches_all
        st.session_state.ap = {
            "date_range": (d_from, d_to),
            "branches": draft_branches, "points": draft_points,
            "groups": draft_groups, "categories": draft_cats,
            "subcategories": draft_subs, "items": draft_items,
            "abc_metric": draft_metric,
        }
        st.rerun()

# Кнопка сброса
if st.sidebar.button("🔄 Сбросить фильтры", use_container_width=True):
    st.session_state.ap = {
        "date_range": (min_d, max_d), "branches": branches_all,
        "points": [], "groups": [], "categories": [], "subcategories": [],
        "items": [], "abc_metric": "Сумма",
    }
    st.session_state.fv += 1
    st.rerun()

# =============================================================================
# Применение фильтров
# =============================================================================
ap = st.session_state.ap
d_from, d_to = ap["date_range"]
from_ts = pd.Timestamp(d_from)
to_ts   = pd.Timestamp(d_to)

df_f = df[(df["Дата"] >= from_ts) & (df["Дата"] <= to_ts)].copy()
df_f = df_f[df_f["Филиал"].isin(ap["branches"])]
if ap["points"]:        df_f = df_f[df_f["Точки"].isin(ap["points"])]
if ap["groups"]:        df_f = df_f[df_f["Группа"].isin(ap["groups"])]
if ap["categories"]:    df_f = df_f[df_f["Категория"].isin(ap["categories"])]
if ap["subcategories"]: df_f = df_f[df_f["Подкатегория"].isin(ap["subcategories"])]
if ap["items"]:         df_f = df_f[df_f["Номенклатура"].isin(ap["items"])]

metric      = ap["abc_metric"]
metric_col  = "Сумма" if metric == "Сумма" else "Количество"
checks_col  = "Склад/Товар"
days_cnt    = max((d_to - d_from).days + 1, 1)

# =============================================================================
# Утилиты форматирования / Excel-экспорт
# =============================================================================
def money(x: float) -> str:
    if pd.isna(x): return "—"
    return f"{x:,.0f}".replace(",", " ")

def num(x: float, decimals: int = 0) -> str:
    if pd.isna(x): return "—"
    fmt = f"{{:,.{decimals}f}}"
    return fmt.format(x).replace(",", " ")

def safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0


# Excel helpers (сохранены из v1)
HEADER_COLOR, HEADER_FONT, ROW_ALT = "1F4E79", "FFFFFF", "DCE6F1"
MONEY_COLS  = {"Выручка","Сумма","Средний чек","Value","PeakValue",
               "Итого за период","Итого (сом)","Остаток (сом)","Прогноз (сом)"}
INT_COLS    = {"Количество","Чеков","SKU_count","Итого (кол)","Остаток (шт)","Прогноз (шт)"}
PCT_COLS    = {"Share","CumShare","SKU_share","Value_share","Доля выручки","Δ %"}
FLOAT2_COLS = {"Позиции/чек","Товаров/чек","Среднее/день (кол)","Среднее/день (сом)"}

def _fmt(col: str) -> str | None:
    if col in MONEY_COLS:  return '#,##0'
    if col in INT_COLS:    return '#,##0'
    if col in PCT_COLS:    return '0.00%'
    if col in FLOAT2_COLS: return '0.00'
    return None

def _autowidth(ws):
    for col in ws.columns:
        ml = 0
        letter = get_column_letter(col[0].column)
        for c in col:
            v = str(c.value) if c.value is not None else ""
            ml = max(ml, len(v))
        ws.column_dimensions[letter].width = min(max(ml + 3, 8), 55)

def df_to_sheet(ws, df_in: pd.DataFrame, title: str | None = None):
    start = 1
    if title:
        tc = ws.cell(row=1, column=1, value=title)
        tc.font = Font(bold=True, size=12, color=HEADER_COLOR)
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(len(df_in.columns), 1))
        start = 2
    fill = PatternFill("solid", fgColor=HEADER_COLOR)
    f_font = Font(bold=True, color=HEADER_FONT, size=10)
    cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, col in enumerate(df_in.columns, 1):
        c = ws.cell(row=start, column=i, value=str(col))
        c.fill, c.font, c.alignment = fill, f_font, cen
    ws.row_dimensions[start].height = 28
    alt = PatternFill("solid", fgColor=ROW_ALT)
    bd = Border(bottom=Side(style="thin", color="BFBFBF"))
    for ri, row in enumerate(df_in.itertuples(index=False), 1):
        is_alt = ri % 2 == 0
        for ci, (col, v) in enumerate(zip(df_in.columns, row), 1):
            c = ws.cell(row=start + ri, column=ci)
            if isinstance(v, pd.Timestamp):
                c.value = v.to_pydatetime(); c.number_format = 'DD.MM.YYYY'
            elif hasattr(v, "item"):
                c.value = v.item()
            else:
                c.value = v
            fm = _fmt(col)
            if fm: c.number_format = fm
            c.font = Font(size=10); c.border = bd
            if is_alt: c.fill = alt
    _autowidth(ws)
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

def build_xlsx_bytes(sheets: list[tuple[str, pd.DataFrame, str | None]]) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for name, dfx, title in sheets:
        if dfx is None or dfx.empty: continue
        safe = name[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=safe)
        df_to_sheet(ws, dfx.reset_index(drop=True), title)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def dl_btn(label, sheets, filename="report.xlsx", key=None):
    clean = [(n, d, t) for n, d, t in sheets if d is not None and not d.empty]
    if not clean: return
    st.download_button(
        f"⬇️ {label}", data=build_xlsx_bytes(clean), file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


# =============================================================================
# Базовые агрегаты (KPI, ABC, время)
# =============================================================================
def count_checks(frame: pd.DataFrame) -> int:
    if checks_col not in frame.columns or frame.empty: return 0
    s = (frame[checks_col].astype(str).str.strip()
         .str.replace(r"\s+", " ", regex=True)
         .replace({"": pd.NA, "nan": pd.NA}).dropna())
    return int(s.nunique())

def kpi_group(frame: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=group_cols + ["Выручка","Количество","Чеков",
                                                   "Средний чек","Позиции/чек","Товаров/чек","Доля выручки"])
    g = frame.groupby(group_cols, dropna=False).agg(
        Выручка=("Сумма","sum"), Количество=("Количество","sum"), Строк=("Сумма","size"),
    ).reset_index()
    checks = frame[group_cols + [checks_col]].copy()
    checks[checks_col] = (checks[checks_col].astype(str).str.strip()
                          .str.replace(r"\s+"," ", regex=True)
                          .replace({"": pd.NA, "nan": pd.NA}))
    checks = checks.dropna(subset=[checks_col])
    if not checks.empty:
        cnt = checks.groupby(group_cols)[checks_col].nunique().reset_index().rename(columns={checks_col: "Чеков"})
        g = g.merge(cnt, on=group_cols, how="left")
    else:
        g["Чеков"] = 0
    g["Чеков"] = g["Чеков"].fillna(0).astype(int)
    g["Средний чек"] = g.apply(lambda r: r["Выручка"]/r["Чеков"] if r["Чеков"] else 0.0, axis=1)
    g["Позиции/чек"] = g.apply(lambda r: r["Строк"]/r["Чеков"] if r["Чеков"] else 0.0, axis=1)
    g["Товаров/чек"] = g.apply(lambda r: r["Количество"]/r["Чеков"] if r["Чеков"] else 0.0, axis=1)
    total = float(g["Выручка"].sum()) or 0.0
    g["Доля выручки"] = g["Выручка"]/total if total else 0.0
    return g.sort_values("Выручка", ascending=False).drop(columns=["Строк"])


kpi_branch       = kpi_group(df_f, ["Филиал"])
kpi_branch_point = kpi_group(df_f, ["Филиал","Точки"])

# ABC
A_THR, B_THR = 0.80, 0.95

def build_abc(frame, group_cols):
    g = (frame.groupby(group_cols, dropna=False)[metric_col]
         .sum().reset_index().rename(columns={metric_col: "Value"})
         .sort_values("Value", ascending=False).reset_index(drop=True))
    total = float(g["Value"].sum()) if not g.empty else 0.0
    if total <= 0 or pd.isna(total):
        g["Share"] = g["CumShare"] = 0.0
    else:
        g["Share"] = g["Value"]/total; g["CumShare"] = g["Share"].cumsum()
    g["ABC"] = g["CumShare"].apply(lambda x: "A" if x <= A_THR else ("B" if x <= B_THR else "C"))
    return g

def abc_summary(abc_df):
    if abc_df.empty:
        return pd.DataFrame(columns=["ABC","SKU_count","SKU_share","Value","Value_share"])
    tsku = len(abc_df); tval = float(abc_df["Value"].sum()) or 0.0
    s = abc_df.groupby("ABC")["Value"].agg(SKU_count="count", Value="sum").reset_index()
    s["SKU_share"] = s["SKU_count"]/tsku if tsku else 0.0
    s["Value_share"] = s["Value"]/tval if tval else 0.0
    s["ABC"] = pd.Categorical(s["ABC"], categories=["A","B","C"], ordered=True)
    return s.sort_values("ABC").reset_index(drop=True)

abc_overall   = build_abc(df_f, ["Номенклатура"])
abc_by_branch = build_abc(df_f, ["Филиал","Номенклатура"])
abc_stats     = abc_summary(abc_overall)

# Time
df_time = df_f.copy()
if "Время" in df_time.columns:
    t = pd.to_datetime(df_time["Время"].astype(str).str.strip(), errors="coerce")
    df_time["Hour"] = t.dt.hour
    df_time = df_time[df_time["Hour"].notna()].copy()
    df_time["Hour"] = df_time["Hour"].astype(int)
    df_time["DOW"]  = df_time["Дата"].dt.dayofweek
else:
    df_time = df_time.iloc[0:0].copy()

# Daily series для текущего периода
df_daily = (df_f.set_index("Дата")[metric_col].resample("D").sum().reset_index()
            .rename(columns={"Дата":"Day", metric_col:"Value"}))


# =============================================================================
# Сравнение периодов (MoM / YoY / WoW)
# =============================================================================
def period_compare(df_all: pd.DataFrame, d_from: date, d_to: date,
                   filters_ap: dict) -> dict:
    """
    Для текущего периода вычисляет предыдущий такого же размера (prev),
    Week-over-Week (−7 дней), Month-over-Month (−30), Year-over-Year (−365).
    Возвращает dict с суммами и % изменениями.
    """
    # применяем все не-датовые фильтры
    base = df_all.copy()
    for k, col in [("branches","Филиал"),("points","Точки"),
                   ("groups","Группа"),("categories","Категория"),
                   ("subcategories","Подкатегория"),("items","Номенклатура")]:
        vals = filters_ap.get(k, [])
        if vals: base = base[base[col].isin(vals)]

    ndays = (d_to - d_from).days + 1

    def _period_sum(p_from, p_to, metric):
        f = base[(base["Дата"] >= pd.Timestamp(p_from)) & (base["Дата"] <= pd.Timestamp(p_to))]
        return float(f[metric].sum()) if not f.empty else 0.0

    cur_sum = _period_sum(d_from, d_to, "Сумма")
    cur_qty = _period_sum(d_from, d_to, "Количество")

    shifts = {
        "prev": ndays,                # предыдущий такой же
        "wow":  7,
        "mom":  30,
        "yoy":  365,
    }
    out = {"cur": {"Сумма": cur_sum, "Количество": cur_qty}}
    for name, shift in shifts.items():
        p_from = d_from - timedelta(days=shift)
        p_to   = d_to   - timedelta(days=shift)
        s = _period_sum(p_from, p_to, "Сумма")
        q = _period_sum(p_from, p_to, "Количество")
        out[name] = {
            "from": p_from, "to": p_to,
            "Сумма": s, "Количество": q,
            "Δ_сумма": (cur_sum - s)/s*100 if s else None,
            "Δ_кол":   (cur_qty - q)/q*100 if q else None,
        }
    return out


# =============================================================================
# Cross-sell (ассоциации внутри чека)
# =============================================================================
@st.cache_data(show_spinner=False)
def build_crosssell(df_in_hash: pd.DataFrame, min_support: int = 5) -> pd.DataFrame:
    """
    Возвращает пары номенклатур, которые чаще всего встречаются в одном чеке.
    Колонки: A, B, Pair_count, Support_A, Support_B, Confidence_A→B, Lift.
    """
    if df_in_hash.empty or checks_col not in df_in_hash.columns:
        return pd.DataFrame()

    # чек -> set позиций
    df_c = df_in_hash[[checks_col, "Номенклатура"]].copy()
    df_c[checks_col] = (df_c[checks_col].astype(str).str.strip()
                        .str.replace(r"\s+"," ", regex=True)
                        .replace({"": pd.NA, "nan": pd.NA}))
    df_c = df_c.dropna(subset=[checks_col, "Номенклатура"])
    # уникальные позиции в чеке
    checks_grouped = df_c.groupby(checks_col)["Номенклатура"].apply(lambda s: sorted(set(s)))
    total_checks = len(checks_grouped)
    if total_checks == 0: return pd.DataFrame()

    item_count: Counter = Counter()
    pair_count: Counter = Counter()
    for items in checks_grouped.values:
        if len(items) < 2: continue
        for it in items: item_count[it] += 1
        for a, b in combinations(items, 2):
            pair_count[(a, b)] += 1

    rows = []
    for (a, b), cnt in pair_count.items():
        if cnt < min_support: continue
        sa = item_count[a]; sb = item_count[b]
        support_a = sa / total_checks
        support_b = sb / total_checks
        conf_ab   = cnt / sa if sa else 0.0   # P(B|A)
        conf_ba   = cnt / sb if sb else 0.0
        lift      = conf_ab / support_b if support_b else 0.0
        rows.append({
            "A": a, "B": b, "Pair_count": cnt,
            "Support_A": support_a, "Support_B": support_b,
            "Confidence_A→B": conf_ab, "Confidence_B→A": conf_ba,
            "Lift": round(lift, 3),
        })
    res = pd.DataFrame(rows)
    if res.empty: return res
    return res.sort_values(["Lift","Pair_count"], ascending=[False,False]).reset_index(drop=True)


# =============================================================================
# Прогноз спроса
# =============================================================================
def forecast_demand(df_in: pd.DataFrame, horizon_days: int,
                    metric: str = "Количество") -> pd.DataFrame:
    """
    Прогноз на horizon_days вперёд.
    Алгоритм: среднее по дню недели за последние 4 недели, с добавлением
    общего тренда (slope от 30 последних дней). Прост, но ощутимо лучше
    плоского среднего.
    """
    if df_in.empty: return pd.DataFrame()
    daily = df_in.groupby(df_in["Дата"].dt.date)[metric].sum()
    if len(daily) < 7:
        avg = float(daily.mean()) if len(daily) else 0.0
        last = pd.Timestamp(daily.index[-1]) if len(daily) else pd.Timestamp.today()
        future = [last + pd.Timedelta(days=i+1) for i in range(horizon_days)]
        return pd.DataFrame({"Дата": future, "Прогноз": [avg]*horizon_days})

    idx = pd.to_datetime(daily.index)
    daily.index = idx
    # день недели: 0..6
    dow_avg = {}
    # последние 4 полных недели
    last28 = daily.last("28D") if len(daily) >= 28 else daily
    for d in range(7):
        vals = last28[last28.index.dayofweek == d]
        dow_avg[d] = float(vals.mean()) if len(vals) else float(daily.mean())

    # линейный тренд по последним 30 дням
    last30 = daily.last("30D") if len(daily) >= 30 else daily
    x = np.arange(len(last30))
    y = last30.values.astype(float)
    if len(x) >= 2 and np.std(x) > 0:
        slope, intercept = np.polyfit(x, y, 1)
    else:
        slope = 0.0

    last_ts = idx.max()
    future_rows = []
    for i in range(1, horizon_days + 1):
        ts = last_ts + pd.Timedelta(days=i)
        base_val = dow_avg[ts.dayofweek]
        # применяем тренд (умеренно)
        adj = base_val + slope * i * 0.3
        future_rows.append({"Дата": ts, "Прогноз": max(adj, 0.0),
                            "День недели": ts.strftime("%a")})
    return pd.DataFrame(future_rows)


# =============================================================================
# Детект аномалий
# =============================================================================
def detect_anomalies(daily_df: pd.DataFrame, z_threshold: float = 2.0) -> pd.DataFrame:
    """
    daily_df: DataFrame с колонками Day, Value.
    Возвращает строки где |Z-score| > threshold.
    """
    if daily_df.empty or len(daily_df) < 7: return pd.DataFrame()
    d = daily_df.copy()
    mu = d["Value"].mean(); sd = d["Value"].std() or 1.0
    d["Z"] = (d["Value"] - mu) / sd
    d["Тип"] = d["Z"].apply(lambda z: "⬆️ Пик" if z > z_threshold else ("⬇️ Провал" if z < -z_threshold else None))
    return d[d["Тип"].notna()].sort_values("Z", ascending=False).reset_index(drop=True)


# =============================================================================
# Поиск SKU — индекс по нижнему регистру
# =============================================================================
@st.cache_data(show_spinner=False)
def sku_index(df_all: pd.DataFrame) -> list[str]:
    return sorted(df_all["Номенклатура"].dropna().astype(str).unique().tolist())


# =============================================================================
# Сезонность: месяц × группа
# =============================================================================
@st.cache_data(show_spinner=False)
def seasonality_matrix(df_all: pd.DataFrame, metric: str = "Сумма") -> pd.DataFrame:
    """Матрица: строки = группы, колонки = 1..12 (месяц), ячейки = доля каждой группы в месяце."""
    if df_all.empty: return pd.DataFrame()
    d = df_all.copy()
    d["Month"] = d["Дата"].dt.month
    pv = d.pivot_table(index="Группа", columns="Month", values=metric, aggfunc="sum", fill_value=0)
    # нормируем по месяцу (доля группы от месячной выручки)
    col_sums = pv.sum(axis=0).replace(0, 1)
    pv_norm = pv.div(col_sums, axis=1)
    return pv_norm


# =============================================================================
# Безопасный сток (как в v1) + BOM
# =============================================================================
SEASONAL_KEYWORDS = ['23 февраля','14 февраля','8 марта','1 сентября',
                     'рамадан','новый год','пасх','наурыз','весна']

@st.cache_data(show_spinner=False)
def load_bom() -> pd.DataFrame:
    # 1) пробуем GitHub Pages, 2) локально
    for src_fn in [
        lambda: load_from_github_pages(GH_PAGES_BOM),
        lambda: (Path(__file__).parent.parent / BOM_NAME).read_bytes(),
        lambda: (Path(__file__).parent / BOM_NAME).read_bytes(),
    ]:
        try:
            bts = src_fn()
            df_bom = pd.read_excel(io.BytesIO(bts))
            df_bom.columns = ["Набор","Компонент","Кол"]
            df_bom["Набор"]     = df_bom["Набор"].astype(str).str.strip()
            df_bom["Компонент"] = df_bom["Компонент"].astype(str).str.strip()
            df_bom["Кол"]       = pd.to_numeric(df_bom["Кол"], errors="coerce").fillna(0)
            df_bom = df_bom[~df_bom["Набор"].str.lower().apply(
                lambda n: any(k in n for k in SEASONAL_KEYWORDS))].reset_index(drop=True)
            return df_bom
        except Exception:
            continue
    return pd.DataFrame(columns=["Набор","Компонент","Кол"])


def build_safety_stock(df_in: pd.DataFrame, days_in_period: int, cover_days: int) -> pd.DataFrame:
    if df_in.empty: return pd.DataFrame()
    extras = [c for c in ["Категория","Подкатегория","Группа"] if c in df_in.columns]
    g = (df_in.groupby("Номенклатура", dropna=False)
         .agg(Итого_кол=("Количество","sum"), Итого_сом=("Сумма","sum"))
         .reset_index())
    g["Среднее/день (кол)"] = g["Итого_кол"]/days_in_period
    g["Среднее/день (сом)"] = g["Итого_сом"]/days_in_period
    g["Остаток (шт)"] = (g["Среднее/день (кол)"]*cover_days).apply(lambda x: max(int(np.ceil(x)),1))
    g["Остаток (сом)"] = g["Среднее/день (сом)"]*cover_days
    if not abc_overall.empty:
        abc_map = abc_overall.set_index("Номенклатура")["ABC"].to_dict()
        g["ABC"] = g["Номенклатура"].map(abc_map).fillna("—")
    else:
        g["ABC"] = "—"
    if extras:
        meta = df_in[["Номенклатура"] + extras].drop_duplicates("Номенклатура").reset_index(drop=True)
        g = g.merge(meta, on="Номенклатура", how="left")
    g = g.rename(columns={"Итого_кол":"Итого (кол)", "Итого_сом":"Итого (сом)"})
    order = ["Номенклатура"] + extras + ["Итого (кол)","Итого (сом)",
             "Среднее/день (кол)","Среднее/день (сом)","Остаток (шт)","Остаток (сом)","ABC"]
    g = g[[c for c in order if c in g.columns]]
    return g.sort_values("Остаток (шт)", ascending=False).reset_index(drop=True)


def build_components(df_sales: pd.DataFrame, bom: pd.DataFrame,
                     days: int, cover: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df_sales.empty or bom.empty:
        return pd.DataFrame(), pd.DataFrame()
    sales_agg = (df_sales.groupby("Номенклатура", dropna=False)["Количество"]
                 .sum().reset_index().rename(columns={"Количество":"Прямые продажи"}))
    sales_map = dict(zip(sales_agg["Номенклатура"], sales_agg["Прямые продажи"]))
    sets_in_bom = set(bom["Набор"].unique())
    sets_sold = {k: v for k, v in sales_map.items() if k in sets_in_bom}
    direct = {k: v for k, v in sales_map.items() if k not in sets_in_bom}

    from_sets: dict[str, float] = {}
    for _, row in bom.iterrows():
        nb, comp, cnt = row["Набор"], row["Компонент"], float(row["Кол"])
        sold = sets_sold.get(nb, 0)
        from_sets[comp] = from_sets.get(comp, 0) + sold * cnt

    comp_unit = {}
    for _, row in bom.iterrows():
        c = row["Компонент"]
        if c not in comp_unit:
            comp_unit[c] = "кг" if float(row["Кол"]) < 1 else "шт"

    all_components = set(direct) | set(from_sets)
    rows_c = []
    for comp in sorted(all_components):
        dq = direct.get(comp, 0); fs = from_sets.get(comp, 0); tot = dq + fs
        unit = comp_unit.get(comp, "шт"); avg = tot/days if days else 0
        stock = (round(avg*cover, 2) if unit == "кг"
                 else (max(int(np.ceil(avg*cover)), 1) if tot > 0 else 0))
        rows_c.append({
            "Компонент": comp, "Ед": unit,
            "Прямые продажи": round(dq,3) if unit=="кг" else int(dq),
            "Из наборов":     round(fs,3) if unit=="кг" else int(fs),
            "Итого":          round(tot,3) if unit=="кг" else int(tot),
            "Среднее/день":   round(avg,3),
            f"Остаток ({unit})": stock,
        })
    df_c = pd.DataFrame(rows_c)
    if not df_c.empty:
        df_c = df_c.sort_values(df_c.columns[-1], ascending=False).reset_index(drop=True)

    rows_s = []
    for nb in sorted(sets_in_bom):
        sold = sets_sold.get(nb, 0); avg = sold/days if days else 0
        rows_s.append({
            "Набор": nb, "Продано (шт)": int(sold),
            "Среднее/день": round(avg, 2),
            "Остаток (коробок)": max(int(np.ceil(avg*cover)), 1) if sold > 0 else 0,
        })
    df_s = pd.DataFrame(rows_s).sort_values("Продано (шт)", ascending=False).reset_index(drop=True)
    return df_c, df_s


# =============================================================================
# ===============================  UI  =========================================
# =============================================================================
min_year = int(df_f["Дата"].dt.year.min()) if not df_f.empty else pd.Timestamp.today().year
max_year = int(df_f["Дата"].dt.year.max()) if not df_f.empty else pd.Timestamp.today().year

st.title(f"📊 Sales Dashboard v2 — {min_year}–{max_year}")
st.caption(
    f"🗓 {d_from:%d.%m.%Y} — {d_to:%d.%m.%Y} ({days_cnt} дн.)  "
    f"│ Филиалы: {len(ap['branches'])}  │ Точки: {len(ap['points']) or '—'}  "
    f"│ Группы: {len(ap['groups']) or '—'}  │ Метрика ABC: {metric_col}"
)

# === KPI ===
sales  = float(df_f["Сумма"].sum())      if "Сумма"      in df_f.columns else 0.0
qty    = float(df_f["Количество"].sum()) if "Количество" in df_f.columns else 0.0
checks = count_checks(df_f)
avg_ch = safe_div(sales, checks)
main_per_day = safe_div(sales, days_cnt)

k1,k2,k3,k4,k5 = st.columns(5)
k1.metric("Выручка",     money(sales))
k2.metric("Количество",  num(qty))
k3.metric("Чеков",       num(checks))
k4.metric("Средний чек", money(avg_ch))
k5.metric("Выручка / день", money(main_per_day))

st.divider()

# =============================================================================
# TABS
# =============================================================================
TAB_NAMES = [
    "🏠 Главная",
    "📈 Тренд + сравнение",
    "🔝 ABC / Pareto",
    "⏰ Пики времени",
    "🏆 Рейтинг точек",
    "🔍 Карточка товара",
    "🛒 Cross-sell",
    "🎯 Прогноз",
    "📅 Календарь SKU",
    "🗓 Сезонность",
    "🏭 Склад + наборы",
    "📋 План / Факт",
]
tabs = st.tabs(TAB_NAMES)

# =============================================================================
# TAB 1 — 🏠 Главная (автоинсайты)
# =============================================================================
with tabs[0]:
    st.subheader("Главная — ключевые решения на сегодня")

    if df_f.empty:
        st.info("Нет данных по выбранным фильтрам.")
    else:
        # ---- Сравнение с прошлым периодом ----
        pc = period_compare(df, d_from, d_to, ap)
        cur_s, prev = pc["cur"]["Сумма"], pc["prev"]["Сумма"]
        delta_prev = (cur_s - prev)/prev*100 if prev else None

        cA, cB, cC, cD = st.columns(4)
        cA.metric("Текущий период", money(cur_s))
        cB.metric("Предыдущий (−N дней)", money(prev),
                  f"{delta_prev:+.1f}%" if delta_prev is not None else "—")
        wow_s, wow_delta = pc["wow"]["Сумма"], pc["wow"]["Δ_сумма"]
        cC.metric("WoW (−7 дней)", money(wow_s),
                  f"{wow_delta:+.1f}%" if wow_delta is not None else "—")
        yoy_s, yoy_delta = pc["yoy"]["Сумма"], pc["yoy"]["Δ_сумма"]
        cD.metric("YoY (−1 год)", money(yoy_s),
                  f"{yoy_delta:+.1f}%" if yoy_delta is not None else "—")

        st.divider()

        # ---- Автоматические инсайты ----
        st.markdown("#### 🧠 Автоматические инсайты")

        col_left, col_right = st.columns(2)

        # --- LEFT: top risers / fallers ---
        with col_left:
            # сравнение SKU cur vs prev
            prev_from = d_from - timedelta(days=days_cnt)
            prev_to   = d_to   - timedelta(days=days_cnt)
            df_prev = df[(df["Дата"] >= pd.Timestamp(prev_from)) & (df["Дата"] <= pd.Timestamp(prev_to))]
            for k, col in [("branches","Филиал"),("points","Точки"),("groups","Группа"),
                           ("categories","Категория"),("subcategories","Подкатегория"),("items","Номенклатура")]:
                if ap.get(k):
                    df_prev = df_prev[df_prev[col].isin(ap[k])]

            cur_sku = df_f.groupby("Номенклатура")["Сумма"].sum()
            prev_sku = df_prev.groupby("Номенклатура")["Сумма"].sum() if not df_prev.empty else pd.Series(dtype=float)
            skus = pd.DataFrame({"cur": cur_sku, "prev": prev_sku}).fillna(0)
            skus = skus[skus["cur"] + skus["prev"] > 0]
            skus["delta_abs"] = skus["cur"] - skus["prev"]
            skus["delta_pct"] = np.where(skus["prev"] > 0,
                                          (skus["cur"]-skus["prev"])/skus["prev"]*100,
                                          np.where(skus["cur"] > 0, 999.0, 0.0))
            # берём только «заметные» по абсолюту — чтобы не ловить мелочь с 0→1 = +∞%
            threshold = skus["cur"].quantile(0.75) if len(skus) > 4 else 0
            significant = skus[(skus["cur"] >= threshold) | (skus["prev"] >= threshold)]

            st.markdown("**📈 Топ-5 растущих SKU** (vs пред. период)")
            risers = significant.sort_values("delta_abs", ascending=False).head(5).reset_index()
            if not risers.empty:
                for _, r in risers.iterrows():
                    pct = f"+{r['delta_pct']:.0f}%" if r['delta_pct'] < 900 else "новый"
                    st.markdown(
                        f"<div class='insight-card ok'>"
                        f"<div class='insight-title'>🟢 {r['Номенклатура']}</div>"
                        f"<div class='insight-body'>+{money(r['delta_abs'])} сом "
                        f"<b>({pct})</b> &nbsp; {money(r['prev'])} → {money(r['cur'])}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("— нет данных для сравнения")

            st.markdown("**📉 Топ-5 падающих SKU**")
            fallers = significant.sort_values("delta_abs", ascending=True).head(5).reset_index()
            fallers = fallers[fallers["delta_abs"] < 0]
            if not fallers.empty:
                for _, r in fallers.iterrows():
                    pct = f"{r['delta_pct']:.0f}%"
                    st.markdown(
                        f"<div class='insight-card danger'>"
                        f"<div class='insight-title'>🔴 {r['Номенклатура']}</div>"
                        f"<div class='insight-body'>{money(r['delta_abs'])} сом "
                        f"<b>({pct})</b> &nbsp; {money(r['prev'])} → {money(r['cur'])}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("— нет падающих SKU")

        # --- RIGHT: anomalies, dead stock, top branches ---
        with col_right:
            st.markdown("**⚠️ Аномальные дни** (|Z| > 2σ)")
            anom = detect_anomalies(df_daily[["Day","Value"]].rename(columns={"Day":"Day","Value":"Value"}))
            if not anom.empty:
                for _, r in anom.head(5).iterrows():
                    day_str = pd.Timestamp(r['Day']).strftime('%d.%m.%Y') if not pd.isna(r['Day']) else "—"
                    kind = "warn" if r["Тип"].startswith("⬆") else "danger"
                    st.markdown(
                        f"<div class='insight-card {kind}'>"
                        f"<div class='insight-title'>{r['Тип']} {day_str}</div>"
                        f"<div class='insight-body'>{money(r['Value'])} "
                        f"(Z = {r['Z']:+.2f})</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("— аномалий не обнаружено")

            st.markdown("**💤 Мёртвые SKU** (не продавались 14+ дней)")
            last_sale = df.groupby("Номенклатура")["Дата"].max()
            # берём только SKU которые раньше продавались в этих фильтрах
            active_skus = set(df_f["Номенклатура"].unique())
            all_active = df[df["Номенклатура"].isin(active_skus)]
            last_sale_fmt = all_active.groupby("Номенклатура")["Дата"].max()
            today = pd.Timestamp(d_to)
            dead = last_sale_fmt[(today - last_sale_fmt).dt.days > 14]
            dead_df = pd.DataFrame({"Последняя продажа": dead}).sort_values("Последняя продажа")
            dead_df["Дней назад"] = (today - dead_df["Последняя продажа"]).dt.days
            if not dead_df.empty:
                for name, r in dead_df.head(5).iterrows():
                    st.markdown(
                        f"<div class='insight-card warn'>"
                        f"<div class='insight-title'>💤 {name}</div>"
                        f"<div class='insight-body'>Последняя продажа "
                        f"{r['Последняя продажа']:%d.%m.%Y} "
                        f"<b>({r['Дней назад']} дней назад)</b></div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("— все SKU активны")

            st.markdown("**🏆 Топ-3 филиала**")
            if not kpi_branch.empty:
                top_br = kpi_branch.head(3)
                for i, (_, r) in enumerate(top_br.iterrows()):
                    medal = ["🥇","🥈","🥉"][i] if i < 3 else "•"
                    st.markdown(
                        f"<div class='insight-card ok'>"
                        f"<div class='insight-title'>{medal} {r['Филиал']}</div>"
                        f"<div class='insight-body'>Выручка: <b>{money(r['Выручка'])}</b> сом "
                        f"&nbsp;│ Чеков: {int(r['Чеков']):,}".replace(",", " ") +
                        f" &nbsp;│ Ср. чек: {money(r['Средний чек'])}"
                        f"</div></div>",
                        unsafe_allow_html=True,
                    )

        st.divider()
        # ---- Мини-тренд ----
        st.markdown("#### 📊 Тренд (текущий период)")
        if not df_daily.empty:
            fig, ax = plt.subplots(figsize=(12, 3))
            ax.fill_between(df_daily["Day"], df_daily["Value"], alpha=0.3, color="#1F4E79")
            ax.plot(df_daily["Day"], df_daily["Value"], color="#1F4E79", linewidth=1.5)
            ax.set_xlabel(""); ax.set_ylabel(metric_col)
            ax.grid(alpha=0.3)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
            fig.autofmt_xdate()
            st.pyplot(fig, clear_figure=True)


# =============================================================================
# TAB 2 — 📈 Тренд + сравнение
# =============================================================================
with tabs[1]:
    st.subheader(f"Тренд и сравнение периодов — метрика: {metric_col}")
    if df_f.empty:
        st.info("Нет данных.")
    else:
        pc = period_compare(df, d_from, d_to, ap)

        st.markdown("#### Сводка сравнения")
        comp_rows = []
        for key, label in [("prev","Предыдущий"),("wow","Неделю назад (−7 дн)"),
                            ("mom","Месяц назад (−30 дн)"),("yoy","Год назад (−365 дн)")]:
            p = pc[key]
            comp_rows.append({
                "Период": label,
                "Даты": f"{p['from']:%d.%m.%Y} — {p['to']:%d.%m.%Y}",
                "Сумма": p["Сумма"],
                "Δ % (сумма)": p["Δ_сумма"] if p["Δ_сумма"] is not None else 0,
                "Количество": p["Количество"],
                "Δ % (кол)":   p["Δ_кол"]   if p["Δ_кол"]   is not None else 0,
            })
        cdf = pd.DataFrame(comp_rows)
        # Текущий на верх
        cdf_display = cdf.copy()
        cdf_display["Сумма"] = cdf_display["Сумма"].apply(money)
        cdf_display["Количество"] = cdf_display["Количество"].apply(lambda v: num(v))
        cdf_display["Δ % (сумма)"] = cdf_display["Δ % (сумма)"].apply(lambda v: f"{v:+.1f}%" if v else "—")
        cdf_display["Δ % (кол)"]   = cdf_display["Δ % (кол)"].apply(lambda v: f"{v:+.1f}%" if v else "—")
        st.dataframe(cdf_display, use_container_width=True, hide_index=True)

        st.divider()

        # График
        freq = "D" if days_cnt <= 62 else ("W-MON" if days_cnt <= 370 else "MS")
        cur_s = df_f.set_index("Дата")[metric_col].resample(freq).sum()

        which = st.multiselect(
            "Показать линии сравнения",
            ["WoW (−7 дн)", "MoM (−30 дн)", "YoY (−1 год)"],
            default=["WoW (−7 дн)", "YoY (−1 год)"],
        )
        fig, ax = plt.subplots(figsize=(13, 4.5))
        ax.plot(cur_s.index, cur_s.values, color="#1F4E79",
                linewidth=2, marker="o", markersize=3, label=f"Текущий ({d_from:%d.%m} — {d_to:%d.%m})")

        palette = {"WoW (−7 дн)":"#E67E22", "MoM (−30 дн)":"#8E44AD", "YoY (−1 год)":"#16A085"}
        for opt in which:
            shift_days = {"WoW (−7 дн)":7, "MoM (−30 дн)":30, "YoY (−1 год)":365}[opt]
            pf = d_from - timedelta(days=shift_days); pt = d_to - timedelta(days=shift_days)
            # строим df с теми же фильтрами (кроме даты)
            dfs_cmp = df[(df["Дата"] >= pd.Timestamp(pf)) & (df["Дата"] <= pd.Timestamp(pt))].copy()
            for k, col in [("branches","Филиал"),("points","Точки"),("groups","Группа"),
                           ("categories","Категория"),("subcategories","Подкатегория"),("items","Номенклатура")]:
                if ap.get(k): dfs_cmp = dfs_cmp[dfs_cmp[col].isin(ap[k])]
            if dfs_cmp.empty: continue
            s_cmp = dfs_cmp.set_index("Дата")[metric_col].resample(freq).sum()
            # сдвигаем индекс обратно в текущий период, чтобы выровнять на графике
            s_cmp_aligned = s_cmp.copy()
            s_cmp_aligned.index = s_cmp_aligned.index + pd.Timedelta(days=shift_days)
            ax.plot(s_cmp_aligned.index, s_cmp_aligned.values,
                    linewidth=1.3, marker="s", markersize=3,
                    color=palette[opt], alpha=0.7, label=opt, linestyle="--")

        ax.set_xlabel(""); ax.set_ylabel(metric_col)
        ax.legend(loc="upper left", framealpha=0.95); ax.grid(alpha=0.3)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        fig.autofmt_xdate()
        st.pyplot(fig, clear_figure=True)

        # Топ-10 дней
        st.markdown("**Топ-10 лучших дней периода**")
        top10 = df_daily.sort_values("Value", ascending=False).head(10).copy()
        top10["День"] = pd.to_datetime(top10["Day"]).dt.strftime("%A")
        top10["Day"] = pd.to_datetime(top10["Day"]).dt.strftime("%d.%m.%Y")
        top10 = top10[["Day","День","Value"]].rename(columns={"Day":"Дата", "Value":metric_col})
        top10[metric_col] = top10[metric_col].apply(money if metric_col=="Сумма" else (lambda x: num(x)))
        st.dataframe(top10, use_container_width=True, hide_index=True)

        dl_btn("Скачать тренд + сравнение",
               [("Сравнение периодов", cdf, "MoM / WoW / YoY"),
                ("Топ-10 дней", df_daily.sort_values("Value", ascending=False).head(10), "Пики периода")],
               filename=f"trend_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_trend")


# =============================================================================
# TAB 3 — 🔝 ABC / Pareto
# =============================================================================
with tabs[2]:
    st.subheader(f"ABC по номенклатуре — метрика: {metric_col}")

    mode = st.radio("Срез", ["Общий","По филиалам"], horizontal=True, key="abc_mode")

    if mode == "Общий":
        c1, c2 = st.columns([1, 2])
        with c1:
            st.markdown("**Сводка A/B/C**")
            st.dataframe(abc_stats, use_container_width=True, hide_index=True)
            top_n = st.slider("Top-N для Pareto", 10, 200, 30, 10, key="abc_topn")
        with c2:
            if abc_overall.empty:
                st.info("Нет данных.")
            else:
                d = abc_overall.head(top_n).copy()
                fig, ax1 = plt.subplots(figsize=(13, 5))
                bars = ax1.bar(range(len(d)), d["Value"], color="#1F4E79", alpha=0.85)
                ax1.set_ylabel(metric_col); ax1.set_xlabel("")
                ax1.set_xticks(range(len(d)))
                ax1.set_xticklabels(d["Номенклатура"].astype(str).tolist(),
                                    rotation=75, ha="right", fontsize=8)
                ax2 = ax1.twinx()
                ax2.plot(range(len(d)), d["CumShare"].values, color="#C0392B", marker="o", linewidth=1.5)
                ax2.set_ylabel("Кум. доля"); ax2.set_ylim(0, 1.05)
                ax2.axhline(A_THR, linestyle="--", color="#27AE60", alpha=0.7, label="A: 80%")
                ax2.axhline(B_THR, linestyle="--", color="#E67E22", alpha=0.7, label="B: 95%")
                ax2.legend(loc="center right")
                ax1.set_title(f"Pareto Top {top_n}")
                fig.tight_layout()
                st.pyplot(fig, clear_figure=True)
        st.markdown("**Таблица ABC**")
        st.dataframe(abc_overall.head(500), use_container_width=True)
        dl_btn("Скачать ABC (общий)",
               [("ABC Общий", abc_overall, f"ABC — {metric_col}"),
                ("ABC Сводка", abc_stats, "Сводка A/B/C")],
               filename=f"abc_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_abc_o")
    else:
        branches_in = sorted(abc_by_branch["Филиал"].dropna().astype(str).unique().tolist())
        if branches_in:
            sel_b = st.selectbox("Филиал", branches_in, key="abc_branch")
            one = abc_by_branch[abc_by_branch["Филиал"].astype(str) == sel_b].copy()
            one = one.sort_values("Value", ascending=False).reset_index(drop=True)
            tot = float(one["Value"].sum()) or 0.0
            if tot > 0:
                one["Share"] = one["Value"]/tot; one["CumShare"] = one["Share"].cumsum()
                one["ABC"] = one["CumShare"].apply(lambda x: "A" if x<=A_THR else ("B" if x<=B_THR else "C"))
            top_n_b = st.slider("Top-N", 10, 200, 30, 10, key="abc_topnb")
            d = one.head(top_n_b)
            if not d.empty:
                fig, ax1 = plt.subplots(figsize=(13, 5))
                ax1.bar(range(len(d)), d["Value"], color="#1F4E79", alpha=0.85)
                ax1.set_xticks(range(len(d)))
                ax1.set_xticklabels(d["Номенклатура"].astype(str).tolist(),
                                    rotation=75, ha="right", fontsize=8)
                ax2 = ax1.twinx()
                ax2.plot(range(len(d)), d["CumShare"].values, color="#C0392B", marker="o")
                ax2.set_ylim(0, 1.05); ax2.axhline(A_THR, linestyle="--", color="#27AE60")
                ax2.axhline(B_THR, linestyle="--", color="#E67E22")
                fig.tight_layout()
                st.pyplot(fig, clear_figure=True)
            st.dataframe(one.head(500), use_container_width=True)
            dl_btn(f"Скачать ABC — {sel_b}",
                   [(f"ABC {sel_b[:20]}", one, f"ABC — {sel_b}")],
                   filename=f"abc_{sel_b}_{d_from:%Y%m%d}.xlsx", key="dl_abc_b")


# =============================================================================
# TAB 4 — ⏰ Пики времени
# =============================================================================
with tabs[3]:
    st.subheader(f"Касса по времени — метрика: {metric_col}")
    if df_time.empty:
        st.info("Нет данных в колонке 'Время'.")
    else:
        # Heatmap: день недели × час
        st.markdown("**Heatmap: день недели × час**")
        hm = df_time.groupby(["DOW","Hour"])[metric_col].sum().reset_index()
        pv = hm.pivot(index="DOW", columns="Hour", values=metric_col).fillna(0)
        dow_names = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
        pv = pv.reindex(range(7))
        fig, ax = plt.subplots(figsize=(13, 3.2))
        im = ax.imshow(pv.values, aspect="auto", cmap="YlOrRd")
        ax.set_yticks(range(7)); ax.set_yticklabels(dow_names)
        ax.set_xticks(range(len(pv.columns)))
        ax.set_xticklabels([f"{int(h):02d}" for h in pv.columns], fontsize=8)
        ax.set_xlabel("Час"); ax.set_title(f"{metric_col} по часам × дням недели")
        plt.colorbar(im, ax=ax, fraction=0.02, pad=0.01)
        st.pyplot(fig, clear_figure=True)

        st.divider()
        c1, c2 = st.columns([1, 2])
        by_bh = (df_time.groupby(["Филиал","Hour"], dropna=False)[metric_col]
                 .sum().reset_index().rename(columns={metric_col:"Value"}))
        peak = (by_bh.sort_values(["Филиал","Value"], ascending=[True, False])
                .groupby("Филиал", as_index=False).head(1)
                .rename(columns={"Hour":"PeakHour","Value":"PeakValue"}))
        pv2 = by_bh.pivot_table(index="Филиал", columns="Hour", values="Value", fill_value=0).sort_index(axis=1)
        with c1:
            st.markdown("**Пиковый час / филиал**")
            pv_disp = peak.copy()
            pv_disp["PeakHour"] = pv_disp["PeakHour"].apply(lambda h: f"{int(h):02d}:00")
            st.dataframe(pv_disp, use_container_width=True, hide_index=True)
        with c2:
            st.markdown("**Филиалы × часы**")
            st.dataframe(pv2, use_container_width=True)

        sel = st.selectbox("График по филиалу", sorted(by_bh["Филиал"].unique()))
        d_bh = by_bh[by_bh["Филиал"]==sel].sort_values("Hour")
        st.line_chart(d_bh.set_index("Hour")["Value"])

        dl_btn("Скачать часы",
               [("Пиковые часы", peak, "Пик по филиалу"),
                ("Филиалы × часы", pv2.reset_index(), f"Матрица часов — {metric_col}"),
                ("По часам детально", by_bh, "Часы × филиалы")],
               filename=f"hours_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_hours")


# =============================================================================
# TAB 5 — 🏆 Рейтинг точек / филиалов
# =============================================================================
with tabs[4]:
    st.subheader("Рейтинг филиалов и точек")
    st.caption("Композитный скор: нормированная выручка + средний чек + позиции/чек.")

    if df_f.empty:
        st.info("Нет данных.")
    else:
        level = st.radio("Уровень", ["Филиалы","Точки (внутри филиалов)"], horizontal=True, key="rk_lvl")
        base_kpi = kpi_branch.copy() if level == "Филиалы" else kpi_branch_point.copy()

        if base_kpi.empty:
            st.info("Нет данных.")
        else:
            # Нормируем три ключевых метрики
            def _normalize(s: pd.Series) -> pd.Series:
                mn, mx = s.min(), s.max()
                return (s - mn) / (mx - mn) if mx > mn else pd.Series([0.5]*len(s), index=s.index)

            base_kpi["n_выручка"]  = _normalize(base_kpi["Выручка"])
            base_kpi["n_ср_чек"]   = _normalize(base_kpi["Средний чек"])
            base_kpi["n_позиций"]  = _normalize(base_kpi["Позиции/чек"])
            # веса: выручка = 50%, ср.чек = 30%, позиции = 20%
            base_kpi["Скор"]       = (base_kpi["n_выручка"]*0.5 +
                                      base_kpi["n_ср_чек"]*0.3 +
                                      base_kpi["n_позиций"]*0.2)
            base_kpi = base_kpi.sort_values("Скор", ascending=False).reset_index(drop=True)
            base_kpi["Ранг"] = base_kpi.index + 1

            show_cols = (["Ранг"] +
                         (["Филиал"] if level == "Филиалы" else ["Филиал","Точки"]) +
                         ["Выручка","Чеков","Средний чек","Позиции/чек","Товаров/чек","Доля выручки","Скор"])
            disp = base_kpi[show_cols].copy()
            disp["Выручка"]      = disp["Выручка"].round(0)
            disp["Средний чек"]  = disp["Средний чек"].round(0)
            disp["Позиции/чек"]  = disp["Позиции/чек"].round(2)
            disp["Товаров/чек"]  = disp["Товаров/чек"].round(2)
            disp["Доля выручки"] = (disp["Доля выручки"]*100).round(1).astype(str) + "%"
            disp["Скор"]         = disp["Скор"].round(3)

            st.dataframe(disp, use_container_width=True, hide_index=True, height=520)

            # Bar chart top-10
            st.markdown("**Скор — топ 10**")
            top10 = base_kpi.head(10).copy()
            fig, ax = plt.subplots(figsize=(11, 4))
            label_col = "Филиал" if level == "Филиалы" else base_kpi.apply(
                lambda r: f"{r['Филиал']} — {r['Точки']}", axis=1).iloc[:10]
            if level == "Филиалы":
                labels = top10["Филиал"].tolist()
            else:
                labels = [f"{r['Филиал']} — {r['Точки']}" for _, r in top10.iterrows()]
            ax.barh(labels[::-1], top10["Скор"][::-1], color="#1F4E79")
            ax.set_xlabel("Скор (0–1)")
            ax.set_xlim(0, 1)
            ax.grid(axis="x", alpha=0.3)
            fig.tight_layout()
            st.pyplot(fig, clear_figure=True)

            dl_btn("Скачать рейтинг",
                   [("Рейтинг", base_kpi.drop(columns=["n_выручка","n_ср_чек","n_позиций"]), "Композитный рейтинг")],
                   filename=f"rating_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_rating")


# =============================================================================
# TAB 6 — 🔍 Карточка товара
# =============================================================================
with tabs[5]:
    st.subheader("Карточка товара")
    st.caption("Введи часть названия — поиск по подстроке (регистро-независимый).")

    all_skus = sku_index(df)
    q = st.text_input("Поиск SKU", placeholder="например: финик, плитка pistachio, капучино")
    if q:
        matches = [s for s in all_skus if q.lower() in s.lower()]
        st.caption(f"Найдено: {len(matches)} позиций")
    else:
        matches = all_skus[:200]

    if not matches:
        st.info("Ничего не найдено.")
    else:
        chosen = st.selectbox("Выбери позицию", matches, key="sku_card_sel")
        df_sku = df_f[df_f["Номенклатура"] == chosen].copy()
        # если отфильтровано по SKU до единицы — всё равно показываем на всю базу
        df_sku_all = df[df["Номенклатура"] == chosen].copy()

        if df_sku.empty:
            st.warning("За выбранный период этот SKU не продавался. Показываю данные по всему диапазону.")
            df_sku = df_sku_all

        # Карточка
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Выручка", money(float(df_sku["Сумма"].sum())))
        c2.metric("Количество", num(float(df_sku["Количество"].sum())))
        c3.metric("Чеков", num(count_checks(df_sku)))
        avg_price = safe_div(float(df_sku["Сумма"].sum()), float(df_sku["Количество"].sum()))
        c4.metric("Ср. цена", money(avg_price))
        days_in_sku = df_sku["Дата"].dt.date.nunique()
        c5.metric("Дней с продажами", num(days_in_sku))

        # Детали
        meta_cols = [c for c in ["Группа","Категория","Подкатегория"] if c in df_sku.columns]
        if meta_cols:
            meta = df_sku[meta_cols].iloc[0]
            st.markdown(
                "<div class='insight-card'>" +
                " &nbsp;│&nbsp; ".join(f"<b>{k}:</b> {v}" for k, v in meta.items()) +
                "</div>", unsafe_allow_html=True)

        st.divider()

        # Тренд по дням
        st.markdown("**Динамика продаж**")
        colA, colB = st.columns(2)
        daily_sku = (df_sku.groupby(df_sku["Дата"].dt.date)
                     .agg(Выручка=("Сумма","sum"), Количество=("Количество","sum"))
                     .reset_index().rename(columns={"Дата":"День"}))
        daily_sku["День"] = pd.to_datetime(daily_sku["День"])
        with colA:
            st.markdown("*По выручке*")
            st.bar_chart(daily_sku.set_index("День")["Выручка"])
        with colB:
            st.markdown("*По количеству*")
            st.bar_chart(daily_sku.set_index("День")["Количество"])

        # По филиалам / точкам
        st.markdown("**Где продаётся**")
        by_br = df_sku.groupby(["Филиал","Точки"]).agg(
            Выручка=("Сумма","sum"), Количество=("Количество","sum"),
            Чеков=(checks_col, "nunique")).reset_index().sort_values("Выручка", ascending=False)
        by_br["Выручка"]    = by_br["Выручка"].apply(money)
        by_br["Количество"] = by_br["Количество"].apply(lambda v: num(v))
        st.dataframe(by_br, use_container_width=True, hide_index=True, height=300)

        # ABC статус
        if chosen in abc_overall["Номенклатура"].values:
            abc_row = abc_overall[abc_overall["Номенклатура"] == chosen].iloc[0]
            st.caption(f"🎯 ABC статус: **{abc_row['ABC']}**  "
                       f"│ Доля: {abc_row['Share']*100:.2f}%  "
                       f"│ Кум. доля: {abc_row['CumShare']*100:.2f}%")

        dl_btn("Скачать карточку",
               [("Динамика", daily_sku, f"Дневная динамика — {chosen[:30]}"),
                ("По точкам", df_sku.groupby(["Филиал","Точки"]).agg(
                    Выручка=("Сумма","sum"), Количество=("Количество","sum")).reset_index(),
                 f"По точкам — {chosen[:30]}")],
               filename=f"sku_{chosen[:20].replace(' ','_')}.xlsx", key="dl_sku")


# =============================================================================
# TAB 7 — 🛒 Cross-sell
# =============================================================================
with tabs[6]:
    st.subheader("Cross-sell — что покупают вместе")
    st.caption("Пары товаров из одного чека. Lift > 1 означает «покупают вместе чаще случайного».")

    if df_f.empty or checks_col not in df_f.columns:
        st.info("Нет данных о чеках.")
    else:
        col1, col2 = st.columns([1, 1])
        with col1:
            min_support = st.slider(
                "Минимум пар в чеке",
                min_value=3, max_value=200, value=10, step=1,
                help="Отсекаем пары, встретившиеся меньше N раз вместе. Выше = надёжнее, но меньше пар.",
            )
        with col2:
            show_top = st.slider("Показывать пар", 10, 200, 40, 10)

        with st.spinner("Считаем ассоциации..."):
            # хэшим df по набору чеков — иначе cache не сработает между фильтрами
            cs = build_crosssell(df_f[[checks_col, "Номенклатура"]].copy(), min_support=min_support)

        if cs.empty:
            st.info(f"Нет пар с частотой ≥ {min_support}. Уменьши порог.")
        else:
            disp = cs.head(show_top).copy()
            disp["Support_A"]         = (disp["Support_A"]*100).round(2).astype(str) + "%"
            disp["Support_B"]         = (disp["Support_B"]*100).round(2).astype(str) + "%"
            disp["Confidence_A→B"]    = (disp["Confidence_A→B"]*100).round(1).astype(str) + "%"
            disp["Confidence_B→A"]    = (disp["Confidence_B→A"]*100).round(1).astype(str) + "%"
            st.dataframe(disp, use_container_width=True, hide_index=True, height=540)

            st.caption(
                "💡 **Как читать:** Если *Confidence A→B = 60%*, значит когда покупают A, "
                "в 60% случаев также берут B. **Lift = 2.5** → пара встречается в 2.5× чаще, "
                "чем ожидаемо при случайности."
            )
            dl_btn("Скачать пары",
                   [("Cross-sell", cs, f"Ассоциации — min_support={min_support}")],
                   filename=f"crosssell_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx", key="dl_cs")


# =============================================================================
# TAB 8 — 🎯 Прогноз
# =============================================================================
with tabs[7]:
    st.subheader("Прогноз спроса")
    st.caption("Алгоритм: среднее по дню недели за последние 4 недели + линейный тренд за 30 дней.")

    if df_f.empty:
        st.info("Нет данных.")
    else:
        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            horizon = st.selectbox("Горизонт", [7, 14, 30], index=0)
        with c2:
            fc_metric = st.radio("Метрика", ["Количество","Сумма"], horizontal=True, key="fc_m")
        with c3:
            level_fc = st.selectbox("Срез", ["Вся выборка","По номенклатуре","По категории","По группе"])

        if level_fc == "Вся выборка":
            fc = forecast_demand(df_f, horizon, metric=fc_metric)
            hist = df_f.groupby(df_f["Дата"].dt.date)[fc_metric].sum().reset_index()
            hist.columns = ["Дата","Факт"]
            hist["Дата"] = pd.to_datetime(hist["Дата"])

            fig, ax = plt.subplots(figsize=(13, 4.5))
            ax.plot(hist["Дата"], hist["Факт"], color="#1F4E79", linewidth=1.5, marker="o", markersize=3, label="Факт")
            if not fc.empty:
                ax.plot(fc["Дата"], fc["Прогноз"], color="#E67E22",
                        linewidth=2, linestyle="--", marker="s", markersize=4, label="Прогноз")
                ax.axvspan(fc["Дата"].min(), fc["Дата"].max(), color="#E67E22", alpha=0.08)
            ax.set_ylabel(fc_metric); ax.legend(); ax.grid(alpha=0.3)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
            fig.autofmt_xdate()
            st.pyplot(fig, clear_figure=True)

            if not fc.empty:
                st.markdown(f"**Прогноз на {horizon} дней**")
                fc_disp = fc.copy()
                fc_disp["Дата"] = fc_disp["Дата"].dt.strftime("%d.%m.%Y (%a)")
                fc_disp["Прогноз"] = fc_disp["Прогноз"].apply(
                    money if fc_metric == "Сумма" else (lambda x: num(x, 1)))
                st.dataframe(fc_disp, use_container_width=True, hide_index=True)
                total_fc = fc["Прогноз"].sum()
                st.metric(f"Итого за {horizon} дн. ({fc_metric})",
                          money(total_fc) if fc_metric == "Сумма" else num(total_fc, 0))

                dl_btn("Скачать прогноз",
                       [("Прогноз", fc, f"Прогноз — {horizon}д — {fc_metric}"),
                        ("Факт", hist, "Исторический факт")],
                       filename=f"forecast_{horizon}d.xlsx", key="dl_fc")
        else:
            col_map = {"По номенклатуре":"Номенклатура","По категории":"Категория","По группе":"Группа"}
            dim = col_map[level_fc]
            top = df_f.groupby(dim)["Сумма"].sum().sort_values(ascending=False).head(20).index.tolist()
            sel = st.selectbox(f"Выбери {dim.lower()}", top, key="fc_sel")
            df_lvl = df_f[df_f[dim] == sel]
            fc = forecast_demand(df_lvl, horizon, metric=fc_metric)
            hist = df_lvl.groupby(df_lvl["Дата"].dt.date)[fc_metric].sum().reset_index()
            hist.columns = ["Дата","Факт"]; hist["Дата"] = pd.to_datetime(hist["Дата"])

            fig, ax = plt.subplots(figsize=(13, 4))
            ax.plot(hist["Дата"], hist["Факт"], color="#1F4E79", linewidth=1.2, marker="o", markersize=2, label="Факт")
            if not fc.empty:
                ax.plot(fc["Дата"], fc["Прогноз"], color="#E67E22",
                        linewidth=2, linestyle="--", marker="s", label="Прогноз")
            ax.set_title(f"{sel} — прогноз на {horizon} дней")
            ax.legend(); ax.grid(alpha=0.3)
            fig.autofmt_xdate()
            st.pyplot(fig, clear_figure=True)

            if not fc.empty:
                total_fc = fc["Прогноз"].sum()
                st.metric(f"Итого за {horizon} дн.",
                          money(total_fc) if fc_metric == "Сумма" else num(total_fc, 0))


# =============================================================================
# TAB 9 — 📅 Календарь SKU (из v1, адаптировано)
# =============================================================================
def _cal_heatmap(daily_series: pd.Series, year: int, month: int,
                 metric_label: str, vmax: float, cmap) -> plt.Figure:
    cal = calendar.monthcalendar(year, month)
    nw = len(cal)
    dow_ru = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
    mo_ru = ["", "Январь","Февраль","Март","Апрель","Май","Июнь",
             "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    fig, ax = plt.subplots(figsize=(7*1.0 + 1.2, (nw+1.2)*1.0))
    ax.set_xlim(0, 7); ax.set_ylim(0, nw+1); ax.axis("off"); ax.set_aspect("equal")
    fig.patch.set_facecolor("#F7F9FC")
    ax.text(3.5, nw+0.65, f"{mo_ru[month]} {year}",
            ha="center", va="center", fontsize=13, fontweight="bold", color="#1F3864")
    for col, dn in enumerate(dow_ru):
        c = "#C0392B" if col >= 5 else "#1F3864"
        ax.text(col+0.5, nw+0.15, dn, ha="center", va="center",
                fontsize=9, fontweight="bold", color=c)
    norm = mcolors.Normalize(vmin=0, vmax=max(vmax, 1))
    for wi, week in enumerate(cal):
        row = nw - 1 - wi
        for dw, dn in enumerate(week):
            x, y = dw, row
            if dn == 0:
                r = mpatches.FancyBboxPatch(
                    (x+0.06, y+0.06), 0.88, 0.88,
                    boxstyle="round,pad=0.04", linewidth=0, facecolor="#ECEFF4")
                ax.add_patch(r); continue
            d = date(year, month, dn)
            val = float(daily_series.get(d, 0.0))
            is_we = dw >= 5
            if val > 0:
                rgba = cmap(norm(val)); face = rgba
                tc = "white" if norm(val) > 0.55 else "#1F3864"
            else:
                face = "#ECEFF4" if not is_we else "#FAE5E5"
                tc = "#AABBCC"
            r = mpatches.FancyBboxPatch(
                (x+0.06, y+0.06), 0.88, 0.88,
                boxstyle="round,pad=0.05", linewidth=0.5,
                edgecolor="#CFD8E3", facecolor=face)
            ax.add_patch(r)
            ax.text(x+0.12, y+0.83, str(dn), ha="left", va="top",
                    fontsize=7, color=tc, alpha=0.7)
            if val > 0:
                s = f"{val:,.1f}".rstrip("0").rstrip(".") if val != int(val) else str(int(val))
                ax.text(x+0.5, y+0.42, s, ha="center", va="center",
                        fontsize=10.5, fontweight="bold", color=tc)
    fig.tight_layout(pad=0.3)
    return fig


def _months_in(d_from, d_to):
    out = []
    cur_d = date(d_from.year, d_from.month, 1)
    end_d = date(d_to.year,   d_to.month,   1)
    while cur_d <= end_d:
        out.append((cur_d.year, cur_d.month))
        cur_d = date(cur_d.year + (1 if cur_d.month==12 else 0),
                     1 if cur_d.month==12 else cur_d.month+1, 1)
    return out


with tabs[8]:
    st.subheader("Календарь SKU — продажи по дням")
    if df_f.empty:
        st.info("Нет данных.")
    else:
        items = sorted(df_f["Номенклатура"].dropna().unique().tolist())
        if not items:
            st.info("Нет SKU.")
        else:
            c1, c2 = st.columns(2)
            with c1:
                cal_metric = st.radio("Метрика", ["Количество","Сумма"],
                                      horizontal=True, key="cal_m")
            with c2:
                show_all = st.checkbox("Все месяцы периода", value=False)

            chosen_item = st.selectbox("Номенклатура", items, key="cal_sku")
            df_it = df_f[df_f["Номенклатура"] == chosen_item]
            if df_it.empty:
                st.info("Нет продаж.")
            else:
                daily = df_it.groupby(df_it["Дата"].dt.date)[cal_metric].sum()
                daily.index = pd.to_datetime(daily.index).date
                total = float(daily.sum()); days = int((daily > 0).sum())
                avg_d = total/days if days else 0.0
                label = "Продано" if cal_metric == "Количество" else "Выручка"
                st.caption(
                    f"📦 {label}: **{money(total) if cal_metric=='Сумма' else num(total,2)}**  "
                    f"│ 📅 Дней с продажами: **{days}**  │ ⌀/день: **"
                    f"{money(avg_d) if cal_metric=='Сумма' else num(avg_d,2)}**"
                )
                cmap = plt.get_cmap("YlOrRd" if cal_metric=="Количество" else "Blues")
                vmax = float(daily.max()) if not daily.empty else 1.0
                months = _months_in(d_from, d_to)
                if not show_all and months:
                    mo_ru = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
                             "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
                    sel_ym = st.selectbox("Месяц", months,
                                          format_func=lambda ym: f"{mo_ru[ym[1]]} {ym[0]}")
                    months = [sel_ym]
                for y, m in months:
                    st.pyplot(_cal_heatmap(daily, y, m, cal_metric, vmax, cmap),
                              clear_figure=True)


# =============================================================================
# TAB 10 — 🗓 Сезонность категорий
# =============================================================================
with tabs[9]:
    st.subheader("Сезонность категорий — по месяцам")
    st.caption("Что продаётся в какой месяц лучше. По доле в месячной выручке.")

    if df.empty:
        st.info("Нет данных.")
    else:
        level_s = st.radio("Группировать по", ["Группа","Категория"], horizontal=True, key="seas_lvl")
        metric_s = st.radio("Метрика", ["Сумма","Количество"], horizontal=True, key="seas_m")

        # Используем весь df, не отфильтрованный — чтобы сезонность была полной
        d = df.copy()
        d["Month"] = d["Дата"].dt.month
        pv = d.pivot_table(index=level_s, columns="Month", values=metric_s, aggfunc="sum", fill_value=0)
        if pv.empty:
            st.info("Нет данных.")
        else:
            # нормируем долю в месяце
            col_sums = pv.sum(axis=0).replace(0, 1)
            pv_norm = pv.div(col_sums, axis=1)

            # Сортировка по суммарной значимости
            pv_norm = pv_norm.loc[pv_norm.sum(axis=1).sort_values(ascending=False).index]
            pv      = pv.loc[pv_norm.index]

            mo_short = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
            pv_disp = pv_norm.copy()
            pv_disp.columns = [mo_short[c-1] for c in pv_disp.columns]

            st.markdown("**Матрица (доля в месячной выручке)**")
            # heatmap
            fig, ax = plt.subplots(figsize=(12, max(4, 0.35*len(pv_disp))))
            im = ax.imshow(pv_disp.values, aspect="auto", cmap="YlGnBu")
            ax.set_xticks(range(len(pv_disp.columns)))
            ax.set_xticklabels(pv_disp.columns)
            ax.set_yticks(range(len(pv_disp)))
            ax.set_yticklabels(pv_disp.index, fontsize=9)
            # подписи %
            for i in range(len(pv_disp)):
                for j in range(len(pv_disp.columns)):
                    v = pv_disp.values[i, j]
                    if v >= 0.05:
                        ax.text(j, i, f"{v*100:.0f}%", ha="center", va="center",
                                fontsize=7, color="white" if v > 0.25 else "#1F3864")
            plt.colorbar(im, ax=ax, fraction=0.015, pad=0.01, label="Доля в месяце")
            fig.tight_layout()
            st.pyplot(fig, clear_figure=True)

            st.markdown("**Абсолютные значения**")
            pv_abs = pv.copy()
            pv_abs.columns = [mo_short[c-1] for c in pv_abs.columns]
            if metric_s == "Сумма":
                st.dataframe(pv_abs.apply(lambda s: s.apply(money)), use_container_width=True)
            else:
                st.dataframe(pv_abs.round(0).astype(int), use_container_width=True)

            dl_btn("Скачать матрицу сезонности",
                   [("Сезонность (доли)", pv_norm.reset_index(), f"Доли по месяцам — {metric_s}"),
                    ("Сезонность (abs)",  pv.reset_index(),      f"Абсолютные — {metric_s}")],
                   filename=f"seasonality_{metric_s.lower()}.xlsx", key="dl_seas")


# =============================================================================
# TAB 11 — 🏭 Склад + наборы
# =============================================================================
with tabs[10]:
    st.subheader("🏭 Неснижаемые остатки — план производства")
    if df_f.empty:
        st.info("Нет данных.")
    else:
        c1, c2, c3 = st.columns(3)
        with c1:
            cover = st.number_input("Норма остатка (дней)", 1, 30, 9, 1,
                                    help="7 дн. до следующего понедельника + 2 дн. буфер = 9.")
        with c2:
            ss_m = st.radio("Метрика", ["Количество","Сумма","Оба"], index=0, horizontal=True)
        with c3:
            abc_f = st.multiselect("ABC", ["A","B","C"], default=["A","B","C"])

        st.caption(
            f"📅 {d_from:%d.%m.%Y}–{d_to:%d.%m.%Y} ({days_cnt} дн.)  "
            f"│ Норма: **{cover} дн.**  │ Формула: `⌀/день × {cover}`"
        )

        ss = build_safety_stock(df_f, days_cnt, cover)
        if ss.empty:
            st.info("Нет данных.")
        else:
            if abc_f and "ABC" in ss.columns:
                ss = ss[ss["ABC"].isin(abc_f + ["—"])]
            base_cols = ["Номенклатура"] + [c for c in ["Группа","Категория","Подкатегория"] if c in ss.columns]
            if ss_m == "Количество":
                show = base_cols + ["Итого (кол)","Среднее/день (кол)","Остаток (шт)","ABC"]
            elif ss_m == "Сумма":
                show = base_cols + ["Итого (сом)","Среднее/день (сом)","Остаток (сом)","ABC"]
            else:
                show = base_cols + ["Итого (кол)","Итого (сом)","Среднее/день (кол)","Среднее/день (сом)",
                                    "Остаток (шт)","Остаток (сом)","ABC"]
            show = [c for c in show if c in ss.columns]
            disp = ss[show].copy()
            for col in ["Среднее/день (кол)","Среднее/день (сом)"]:
                if col in disp.columns: disp[col] = disp[col].round(2)
            for col in ["Итого (сом)","Остаток (сом)"]:
                if col in disp.columns: disp[col] = disp[col].round(0)

            m1, m2, m3 = st.columns(3)
            m1.metric("Позиций", len(disp))
            if "Остаток (шт)" in disp.columns:
                m2.metric("Остаток итого (шт)", num(int(disp["Остаток (шт)"].sum())))
            if "Остаток (сом)" in disp.columns:
                m3.metric("Остаток итого (сом)", money(disp["Остаток (сом)"].sum()))

            st.dataframe(disp, use_container_width=True, hide_index=True, height=520)
            dl_btn("Скачать нормы остатков",
                   [("Нормы остатков", ss, f"Неснижаемые {d_from:%d.%m}-{d_to:%d.%m} норма {cover}д")],
                   filename=f"norms_{d_from:%Y%m%d}_{d_to:%Y%m%d}_{cover}d.xlsx", key="dl_ss")

            # BOM
            st.divider()
            st.markdown("#### 📦 Разбивка наборов")
            bom = load_bom()
            if bom.empty:
                st.info("Файл разбивка_наборов.xlsx не найден.")
            else:
                dfc, dfs = build_components(df_f, bom, days_cnt, cover)
                b1, b2 = st.columns(2)
                with b1:
                    st.markdown("**Компоненты (производство)**")
                    if dfc.empty: st.info("—")
                    else: st.dataframe(dfc, use_container_width=True, hide_index=True, height=400)
                with b2:
                    st.markdown("**Наборы (сборка)**")
                    if dfs.empty: st.info("—")
                    else: st.dataframe(dfs, use_container_width=True, hide_index=True, height=400)
                if not (dfc.empty and dfs.empty):
                    dl_btn("Скачать разбивку",
                           [("Компоненты", dfc, f"Производство {d_from:%d.%m}-{d_to:%d.%m}"),
                            ("Наборы", dfs, f"Сборка наборов {cover}д")],
                           filename=f"breakdown_{d_from:%Y%m%d}_{cover}d.xlsx", key="dl_bom")


# =============================================================================
# TAB 12 — 📋 План / Факт
# =============================================================================
def _plan_block(df_base: pd.DataFrame, idx: int) -> pd.DataFrame | None:
    key = f"pl_{idx}"
    col_d, col_m = st.columns([3, 1])
    with col_d:
        period = st.date_input("Период", value=(d_from, d_to),
                               min_value=min_d, max_value=max_d,
                               format="DD.MM.YYYY", key=f"{key}_d",
                               label_visibility="collapsed")
    with col_m:
        m = st.radio("", ["Сумма","Количество"],
                     index=0 if metric_col == "Сумма" else 1,
                     horizontal=True, key=f"{key}_m", label_visibility="collapsed")
    if not isinstance(period, tuple) or len(period) != 2:
        st.warning("Выбери период."); return None
    pf, pt = period
    if pf > pt: pf, pt = pt, pf
    dfp = df_base[(df_base["Дата"] >= pd.Timestamp(pf)) & (df_base["Дата"] <= pd.Timestamp(pt))]
    if dfp.empty: st.info("Нет данных."); return None
    col = "Сумма" if m == "Сумма" else "Количество"
    days_p = max((pt - pf).days + 1, 1)
    lbl = "Выручка" if col == "Сумма" else "Кол-во"
    g = (dfp.groupby(["Филиал","Точки"], dropna=False)[col]
         .sum().reset_index().rename(columns={col: lbl}))
    g["⌀/день"] = (g[lbl] / days_p).round(1)
    bt = g.groupby("Филиал")[lbl].sum().reset_index().rename(columns={lbl:"BT"})
    g = g.merge(bt, on="Филиал", how="left")
    g["Доля"] = (g[lbl]/g["BT"]*100).round(1).astype(str) + "%"
    g = g.drop(columns=["BT"]).sort_values(["Филиал", lbl], ascending=[True, False]).reset_index(drop=True)
    tr = pd.DataFrame([{
        "Филиал":"ИТОГО","Точки":"—", lbl: g[lbl].sum(),
        "⌀/день": round(g[lbl].sum()/days_p, 1), "Доля":"100%"
    }])
    gd = pd.concat([g, tr], ignore_index=True)
    if col == "Сумма":
        gd[lbl] = gd[lbl].apply(money); gd["⌀/день"] = gd["⌀/день"].apply(money)
    else:
        gd[lbl] = gd[lbl].apply(lambda v: num(v)); gd["⌀/день"] = gd["⌀/день"].apply(lambda v: f"{v:.1f}")
    st.caption(f"📅 {pf:%d.%m.%Y} — {pt:%d.%m.%Y} | {days_p} дн. | {lbl}")
    st.dataframe(gd, use_container_width=True, hide_index=True,
                 height=min(35 * len(gd) + 38, 520))
    exp = g.copy(); exp.insert(0, "Период", f"{pf:%d.%m.%Y}–{pt:%d.%m.%Y}")
    return exp


with tabs[11]:
    st.subheader("План / Факт — сравнение периодов по точкам")
    if df_f.empty:
        st.info("Нет данных.")
    else:
        dfb = df.copy()
        for k, col in [("branches","Филиал"),("points","Точки"),("groups","Группа"),
                       ("categories","Категория"),("subcategories","Подкатегория"),("items","Номенклатура")]:
            if ap.get(k): dfb = dfb[dfb[col].isin(ap[k])]

        if "pl_count" not in st.session_state:
            st.session_state["pl_count"] = 2
        n = st.session_state["pl_count"]

        b1, b2, _ = st.columns([1, 1, 6])
        with b1:
            if st.button("＋ Период", use_container_width=True, key="pl_add"):
                st.session_state["pl_count"] += 1; st.rerun()
        with b2:
            if n > 1 and st.button("－ Убрать", use_container_width=True, key="pl_rem"):
                st.session_state["pl_count"] -= 1; st.rerun()

        st.divider()
        frames = []
        pairs = [(i, i+1) for i in range(0, n, 2)]
        for pair in pairs:
            cs = st.columns(len([p for p in pair if p < n]), gap="large")
            for ci, bi in enumerate(pair):
                if bi >= n: break
                with cs[ci]:
                    st.markdown(
                        f"<div style='background:#1F4E79;border-radius:6px;"
                        f"padding:4px 12px;margin-bottom:8px;'>"
                        f"<span style='color:white;font-weight:700;font-size:13px;'>"
                        f"Период {bi+1}</span></div>", unsafe_allow_html=True)
                    r = _plan_block(dfb, bi)
                    if r is not None: frames.append((f"Период {bi+1}", r))
            if pair != pairs[-1]: st.divider()

        if frames:
            st.divider()
            sheets = [(n, f, f"Факт — {n}") for n, f in frames]
            sheets.append(("Все периоды", pd.concat([f for _, f in frames], ignore_index=True),
                           "Сводная"))
            dl_btn("Скачать план/факт", sheets,
                   filename=f"plan_fact_{d_from:%Y%m%d}.xlsx", key="dl_pl")


# =============================================================================
# Footer
# =============================================================================
st.divider()
st.caption(
    f"© Sales Dashboard v2 │ Данные: GitHub Pages (обновить Excel — в репо `/docs/`) │ "
    f"Маппинг категорий: `category_mapping.py`"
)
