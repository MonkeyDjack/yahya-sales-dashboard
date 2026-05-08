"""
Разовый скрипт: добавляет 6 SKU в Лист1 исходного Итоговый_отчет1.xlsx,
чтобы маппинг сохранился при следующей пересборке.
Делает резервную копию файла перед записью.
"""
from __future__ import annotations
import sys, shutil, time
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

ROOT = Path(__file__).parent.parent
SRC  = ROOT / "Итоговый_отчет1.xlsx"
BAK  = ROOT / "Итоговый_отчет1.bak.xlsx"

# Колонки в Лист1: Номенклатура | Точки | Категория | Под категория
NEW_ROWS = [
    ("Смузи Клубника Банан 450 мл В",            "Бар",     "Смузи",   "нет подкатегорий"),
    ("Латте на миндаль молоке 350 мл В",         "Бар",     "Кофе",    "Латте"),
    ("Флэт Уайт  200 мл  В",                     "Бар",     "Кофе",    "флэт уайт"),
    ("Кофе Бамбл  450 мл на свежевыжатом соке",  "Бар",     "Кофе",    "Бамбл"),
    ("Соленая карамель-Талкан корпусная НК",     "Магазин", "Конфеты", "конфеты корпусные"),
    ("Тарелка постановочная Waseela 25d",        "Магазин", "Другое",  "нет подкатегорий"),
]


def main() -> None:
    if not SRC.exists():
        print(f"❌ Не найден {SRC}")
        return

    print(f"Резервная копия → {BAK} ...")
    shutil.copy2(str(SRC), str(BAK))
    print(f"  размер копии: {BAK.stat().st_size/1024/1024:.1f} МБ")

    print(f"\nОткрытие {SRC} (это займёт время — файл большой) ...")
    t = time.time()
    wb = openpyxl.load_workbook(str(SRC), data_only=False, keep_vba=False)
    print(f"  открыто за {time.time()-t:.1f} с, листы: {wb.sheetnames}")

    ws = wb["Лист1"]
    start_row = ws.max_row + 1
    print(f"\nЛист1: было {ws.max_row} строк, добавляю с строки {start_row}:")

    # Проверяем, что таких номенклатур ещё нет
    existing = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            existing.add(str(v).strip())

    appended = 0
    for row in NEW_ROWS:
        nom = row[0].strip()
        if nom in existing:
            print(f"   ⚠ {nom} — уже есть в Лист1, пропускаю")
            continue
        new_row = ws.max_row + 1
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=new_row, column=col_idx, value=val)
        print(f"   ✓ row {new_row}: {row}")
        appended += 1

    print(f"\nДобавлено строк: {appended}")
    print(f"Сохранение → {SRC} (это тоже займёт время) ...")
    t = time.time()
    wb.save(str(SRC))
    wb.close()
    print(f"  сохранено за {time.time()-t:.1f} с, размер {SRC.stat().st_size/1024/1024:.1f} МБ")
    print("\n✅ Готово")
    print(f"   Резерв сохранён: {BAK}")
    print(f"   При следующем запуске build_refactored_excel.py маппинг подтянется.")


if __name__ == "__main__":
    main()
