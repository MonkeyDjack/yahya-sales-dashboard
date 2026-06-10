"""Excel-экспорт: форматированный (df_to_sheet) и простой (plain_*)."""
from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_COLOR, HEADER_FONT, ROW_ALT = "1F4E79", "FFFFFF", "DCE6F1"
MONEY_COLS  = {"Выручка", "Сумма", "Средний чек", "Value", "PeakValue",
               "Итого за период", "Итого (сом)", "Остаток (сом)", "Прогноз (сом)"}
INT_COLS    = {"Количество", "Чеков", "SKU_count", "Итого (кол)", "Остаток (шт)", "Прогноз (шт)"}
PCT_COLS    = {"Share", "CumShare", "SKU_share", "Value_share", "Доля выручки", "Δ %"}
FLOAT2_COLS = {"Позиции/чек", "Товаров/чек", "Среднее/день (кол)", "Среднее/день (сом)"}


def _fmt(col: str) -> str | None:
    if col in MONEY_COLS:
        return '#,##0'
    if col in INT_COLS:
        return '#,##0'
    if col in PCT_COLS:
        return '0.00%'
    if col in FLOAT2_COLS:
        return '0.00'
    return None


def _autowidth(ws):
    for col in ws.columns:
        ml = 0
        letter = get_column_letter(col[0].column)
        for c in col:
            v = str(c.value) if c.value is not None else ""
            ml = max(ml, len(v))
        ws.column_dimensions[letter].width = min(max(ml + 3, 8), 55)


def _to_excel_safe(v):
    if v is None or v is pd.NA or v is pd.NaT:
        return None
    if isinstance(v, pd.Timestamp):
        return None if pd.isna(v) else v.to_pydatetime()
    if isinstance(v, (datetime, date)):
        return v
    if isinstance(v, (str, int, float, bool)):
        return None if isinstance(v, float) and pd.isna(v) else v
    if hasattr(v, "item"):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return str(v)


def df_to_sheet(ws, df_in: pd.DataFrame, title: str | None = None):
    start = 1
    if title:
        tc = ws.cell(row=1, column=1, value=title)
        tc.font = Font(bold=True, size=12, color=HEADER_COLOR)
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(len(df_in.columns), 1))
        start = 2
    # ВАЖНО: именно start_color/end_color/fill_type — позиционная форма
    # PatternFill("solid", fgColor=...) не пишет applyFill и заливка теряется
    # в старых Excel/LibreOffice (см. CLAUDE.md)
    fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    f_font = Font(bold=True, color=HEADER_FONT, size=10)
    cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, col in enumerate(df_in.columns, 1):
        c = ws.cell(row=start, column=i, value=str(col))
        c.fill, c.font, c.alignment = fill, f_font, cen
    ws.row_dimensions[start].height = 28
    alt = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
    bd = Border(bottom=Side(style="thin", color="BFBFBF"))
    for ri, row in enumerate(df_in.itertuples(index=False), 1):
        is_alt = ri % 2 == 0
        for ci, (col, v) in enumerate(zip(df_in.columns, row), 1):
            c = ws.cell(row=start + ri, column=ci)
            safe = _to_excel_safe(v)
            if isinstance(safe, datetime):
                c.value = safe
                c.number_format = 'DD.MM.YYYY'
            else:
                try:
                    c.value = safe
                except ValueError:
                    c.value = str(safe) if safe is not None else None
            fm = _fmt(col)
            if fm:
                c.number_format = fm
            c.font = Font(size=10)
            c.border = bd
            if is_alt:
                c.fill = alt
    _autowidth(ws)
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


def build_xlsx_bytes(sheets: list[tuple[str, pd.DataFrame, str | None]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, dfx, title in sheets:
        if dfx is None or dfx.empty:
            continue
        safe = name[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=safe)
        df_to_sheet(ws, dfx.reset_index(drop=True), title)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def dl_btn(label, sheets, filename="report.xlsx", key=None):
    clean = [(n, d, t) for n, d, t in sheets if d is not None and not d.empty]
    if not clean:
        return
    st.download_button(
        f"⬇️ {label}", data=build_xlsx_bytes(clean), file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


# --- Простой Excel: без заливок, только жирный хедер + числовые форматы ---
def plain_df_to_sheet(ws, df_in: pd.DataFrame):
    bold = Font(bold=True)
    for i, col in enumerate(df_in.columns, 1):
        ws.cell(row=1, column=i, value=str(col)).font = bold
    for ri, row in enumerate(df_in.itertuples(index=False), 1):
        for ci, (col, v) in enumerate(zip(df_in.columns, row), 1):
            c = ws.cell(row=1 + ri, column=ci)
            safe = _to_excel_safe(v)
            if isinstance(safe, datetime):
                c.value = safe
                c.number_format = 'DD.MM.YYYY'
            else:
                try:
                    c.value = safe
                except ValueError:
                    c.value = str(safe) if safe is not None else None
            fm = _fmt(col)
            if fm:
                c.number_format = fm
    _autowidth(ws)
    ws.freeze_panes = ws.cell(row=2, column=1)


def plain_xlsx_bytes(sheets: list[tuple[str, pd.DataFrame]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, dfx in sheets:
        if dfx is None or dfx.empty:
            continue
        safe = name[:31].replace("/", "-").replace("\\", "-")
        plain_df_to_sheet(wb.create_sheet(title=safe), dfx.reset_index(drop=True))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def dl_btn_plain(label, sheets, filename="report.xlsx", key=None):
    clean = [(n, d) for n, d in sheets if d is not None and not d.empty]
    if not clean:
        st.info("Нет данных для выгрузки за выбранный период.")
        return
    st.download_button(
        f"⬇️ {label}", data=plain_xlsx_bytes(clean), file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )
